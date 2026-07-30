"""Spreadsheet import of already-attended conferences.

The format contract (IMPORT_COLUMNS) is served by /import/format,
renders the UI popup, generates the template, and drives the parser —
one constant, four consumers. These tests lock the round trip: template
out, filled sheet in, conferences + attended participation + outcomes
recorded, idempotent on re-upload.
"""

from __future__ import annotations

import io
import uuid

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import text


def _sheet(rows: list[dict]) -> bytes:
    from app.services.conferences import IMPORT_COLUMNS

    wb = Workbook()
    ws = wb.active
    keys = [c["key"] for c in IMPORT_COLUMNS]
    ws.append(keys)
    for r in rows:
        ws.append([r.get(k, "") for k in keys])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def _mk_sme(engine, name: str) -> str:
    sid = str(uuid.uuid4())
    async with engine.begin() as conn:
        await conn.execute(
            text(
                "INSERT INTO app.smes (id, full_name, team, audience_focus, "
                "location_country, bio, languages, external_links, is_active) "
                "VALUES (:id, :name, 'Eng', '{}', 'US', "
                "'A sufficiently long bio for import testing.', '{}', '{}', true)"
            ),
            {"id": sid, "name": name},
        )
    return sid


@pytest.mark.asyncio
async def test_format_template_and_import_round_trip(
    async_client: AsyncClient, clean_db, test_engine
) -> None:
    fmt = await async_client.get("/api/v1/conferences/import/format")
    assert fmt.status_code == 200
    keys = [c["key"] for c in fmt.json()]
    assert "name" in keys and "attended_by" in keys and "worth_it" in keys

    tpl = await async_client.get("/api/v1/conferences/import/template")
    assert tpl.status_code == 200
    ws = load_workbook(io.BytesIO(tpl.content)).active
    assert [c.value for c in ws[1]] == keys, "template header must match the format contract"

    sme_name = "Import Roster SME"
    await _mk_sme(test_engine, sme_name)

    sheet = _sheet(
        [
            {
                "name": "Imported Summit",
                "start_date": "2024-06-01",
                "country": "US",
                "attended_by": f"{sme_name}; Outside Guest",
                "spend_usd": "1500",
                "leads_generated": "9",
                "worth_it": "yes",
                "notes": "solid event",
            }
        ]
    )
    r = await async_client.post(
        "/api/v1/conferences/import",
        files={"file": ("past.xlsx", sheet, "application/octet-stream")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 1 and body["errors"] == 0

    async with test_engine.begin() as conn:
        conf = (
            await conn.execute(
                text(
                    "SELECT id, status, spend_usd, leads_generated, attendance_verdict "
                    "FROM app.conferences WHERE name='Imported Summit'"
                )
            )
        ).one()
        assert conf.status == "approved"
        assert conf.spend_usd == 1500
        assert conf.leads_generated == 9
        assert conf.attendance_verdict == "would_attend"
        people = (
            await conn.execute(
                text(
                    "SELECT person_label, sme_id IS NOT NULL AS linked, "
                    "attended_at IS NOT NULL AS attended "
                    "FROM app.participation WHERE conference_id=:cid ORDER BY person_label"
                ),
                {"cid": str(conf.id)},
            )
        ).all()
    assert [(p.person_label, p.linked, p.attended) for p in people] == [
        (sme_name, True, True),
        ("Outside Guest", False, True),
    ]

    # Idempotent: same sheet again updates, never duplicates.
    r2 = await async_client.post(
        "/api/v1/conferences/import",
        files={"file": ("past.xlsx", sheet, "application/octet-stream")},
    )
    assert r2.json()["created"] == 0
    assert r2.json()["updated_existing"] == 1
    async with test_engine.begin() as conn:
        n = (
            await conn.execute(
                text(
                    "SELECT count(*) FROM app.participation p "
                    "JOIN app.conferences c ON c.id=p.conference_id "
                    "WHERE c.name='Imported Summit'"
                )
            )
        ).scalar_one()
    assert n == 2


@pytest.mark.asyncio
async def test_import_rejects_junk(async_client: AsyncClient, clean_db) -> None:
    r = await async_client.post(
        "/api/v1/conferences/import",
        files={"file": ("past.txt", b"not a sheet", "text/plain")},
    )
    assert r.status_code == 422

    # A row with no name errors that row without killing the file.
    sheet = _sheet([{"name": "", "city": "Nowhere"}, {"name": "Named Conf"}])
    r2 = await async_client.post(
        "/api/v1/conferences/import",
        files={"file": ("past.xlsx", sheet, "application/octet-stream")},
    )
    body = r2.json()
    assert body["errors"] == 1
    assert body["created"] == 1
