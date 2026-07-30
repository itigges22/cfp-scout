"""The conference itself: its text, its status, its series, who went.

WHAT THIS DOES
    Everything about a conference that is neither a route nor a score.

        text          the blob it gets embedded from
        status        the lifecycle rules — new, planning, attended
        past          have we been to this event before?
        enrichment    an LLM pass that fills gaps in a thin scraped row
        series        "is this the 2027 edition of the same event?"
        participation who is going, what it cost, what came back

HOW IT CONNECTS
    Called by   the conference routes, api/v1/conference_series.py,
                api/v1/participation.py, services/matcher/,
                services/discovery.py, services/extraction.py,
                app/maintenance.py, tasks.py
    Writes      conferences, conference_series, attendance, participation

WORTH KNOWING
    This is the product's spine: a scraped conference is "new", a human
    marks it planning-to-attend, details get filled in, and once it has
    happened the outcome questions open up. Status is a DECISION — no
    background job ever writes it.

    ``series`` was a separate module that imported ``same_series`` from
    this one; the identity rule and the rows it groups are one subject.

    MEASURED, do not remove: ablating the topic lines from the embed blob
    took the corpus from 16 label inversions to 31 (D3, 2026-07-27,
    tests/unit/test_ranking_quality.py).
"""

from __future__ import annotations

import re
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import Final
from uuid import UUID

import structlog
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Conference, ConferenceSeries, Participation, Sme
from app.scheduler import enqueue_task
from app.schemas import (
    AttendanceSummary,
    ParticipationCreate,
    ParticipationRead,
    ParticipationUpdate,
)
from app.services.llm import ChatMessage, ChatRequest, get_llm_client
from app.services.records import model_to_audit_dict, write_audit
from app.settings import get_settings

log = structlog.get_logger("scout.conferences")


# ==========================================================================
# conferences.py
# ==========================================================================


def conference_embed_text(c: Conference) -> str:
    """Compose the descriptive blob we embed for the matcher.

    Preference order for the descriptive body, strongest first:

      1. ``description``           what the event said on its own page
      2. ``enriched_description``  an LLM's guess, from the name
      3. nothing                   name + structural fields alone

    Only reach step 2 when there is no page to have read — a row from a
    name-only feed, or one not yet re-extracted since the description
    column landed. A guess is a reasonable answer to "we have nothing";
    it is the wrong answer to "we have the real thing".
    """
    parts: list[str] = [c.name]
    # Real text beats generated text, always. `description` is what the
    # event said about itself on its own page; `enriched_description` is
    # an LLM's guess made from the name. Preferring the guess when the
    # real thing exists would mean scoring conferences on our own prose.
    if c.description:
        parts.append(c.description)
    elif c.enriched_description:
        parts.append(c.enriched_description)
    # MEASURED, do not remove: ablating the topic lines from this blob
    # took the corpus from 16 label inversions to 31 (D3, 2026-07-27,
    # tests/unit/test_ranking_quality.py). Topic vocabulary is exactly the
    # specific technical wording cosine similarity needs to separate
    # "genuinely about inference" from "AI-adjacent in general".
    if c.topics:
        parts.append("Topics: " + ", ".join(c.topics))
    if c.cfp_topics_of_interest:
        parts.append("CFP topics: " + ", ".join(c.cfp_topics_of_interest))
    if c.location_city or c.location_country:
        loc = " / ".join(p for p in (c.location_city, c.location_country) if p)
        parts.append(f"Location: {loc}")
    if c.is_virtual:
        parts.append("Virtual event.")
    if c.venue:
        parts.append(f"Venue: {c.venue}")
    return "\n".join(parts)


DISCOVERED: Final = "discovered"


NEEDS_REVIEW: Final = "needs_review"


NEEDS_REVIEW_PILLAR: Final = "needs_review_pillar"


NEEDS_SME_REVIEW: Final = "needs_sme_review"


APPROVED: Final = "approved"


REJECTED: Final = "rejected"


VETOED: Final = "vetoed"


LOW_MESSAGING_FIT: Final = "low_messaging_fit"


QUARANTINED: Final = "quarantined"


ALL: Final[frozenset[str]] = frozenset(
    {
        DISCOVERED,
        NEEDS_REVIEW,
        NEEDS_REVIEW_PILLAR,
        NEEDS_SME_REVIEW,
        APPROVED,
        REJECTED,
        VETOED,
        LOW_MESSAGING_FIT,
        QUARANTINED,
    }
)


HIDDEN_FROM_FINDER: Final[frozenset[str]] = frozenset(
    {QUARANTINED, REJECTED}
)


SCOREABLE: Final[frozenset[str]] = ALL - {QUARANTINED}


DIGEST_ELIGIBLE: Final[frozenset[str]] = ALL - HIDDEN_FROM_FINDER - {
    LOW_MESSAGING_FIT,
    VETOED,
}


def is_hidden(status: str) -> bool:
    return status in HIDDEN_FROM_FINDER


def is_scoreable(status: str) -> bool:
    return status in SCOREABLE


def _has_history(conference_name: str, past_name: str) -> str:
    """``"same_event"`` | ``"same_series"`` | ``"unrelated"``."""
    return relationship(conference_name, past_name)


async def load_attended_names(db: AsyncSession) -> tuple[str, ...]:
    """Names of every conference somebody actually went to.

    Attendance is the existence of participation rows, so there is no
    "did we attend" flag that could disagree with the people recorded
    against the event.

    A tuple, not a set: consumers compare by series identity rather than
    by exact lookup. A few dozen strings, cheap to hold for one request.
    """
    rows = (
        (
            await db.execute(
                select(Conference.name)
                .distinct()
                .join(Participation, Participation.conference_id == Conference.id)
            )
        )
        .scalars()
        .all()
    )
    return tuple(n for n in rows if n and n.strip())


def is_previously_attended(conference_name: str, attended_names: tuple[str, ...]) -> bool:
    """True when we have history with this conference or its franchise.

    Deliberately broad: the conference list uses this for the new-vs-
    returning filter, and "we have been to a KubeCon" is useful there even
    when it was a different region.
    """
    if not conference_name or not attended_names:
        return False
    return any(
        _has_history(conference_name, past) != "unrelated" for past in attended_names
    )


def best_verdict_for(
    conference_name: str, attended: Mapping[str, str]
) -> str | None:
    """The verdict of the most closely related past conference, or None.

    ``attended`` maps a past conference name to its verdict. Precedence is
    the relationship strength, not string similarity: a verdict recorded
    against the SAME event beats one from a sibling in the franchise.

    That is what makes per-edition preferences work. "vLLM Meetup Mumbai:
    would_not_attend" must not drag down "vLLM Meetup Boston", and it no
    longer does — those are different events.
    """
    if not conference_name or not attended:
        return None
    best: str | None = None
    best_rank = 0
    for past_name, verdict in attended.items():
        rank = {"same_event": 2, "same_series": 1}.get(
            _has_history(conference_name, past_name), 0
        )
        if rank > best_rank:
            best_rank, best = rank, verdict
    return best


PROMPT_VERSION = "conference.enrichment.v2"




def _build_user_prompt(*, name: str, topics: list[str], country: str | None, city: str | None, is_virtual: bool) -> str:
    location_parts: list[str] = []
    if is_virtual:
        location_parts.append("Virtual")
    elif city or country:
        location_parts.append(", ".join(p for p in (city, country) if p))
    location = location_parts[0] if location_parts else "Location TBD"
    topic_str = ", ".join(topics) if topics else "(no topics tagged)"
    return (
        f"Conference name: {name}\n"
        f"Topics: {topic_str}\n"
        f"Location: {location}\n\n"
        "Write the 2-3 sentence description now."
    )


async def enrich_conference(
    *,
    db: AsyncSession,
    name: str,
    topics: list[str],
    country: str | None = None,
    city: str | None = None,
    is_virtual: bool = False,
) -> str | None:
    """Generate a factual 2-3 sentence description for a conference.

    Returns None on LLM failure — caller can leave the row's
    ``enriched_description`` NULL and the matcher will fall back to the
    bare name+topics text. We don't raise so a single bad row doesn't
    poison a bulk backfill.
    """
    user_prompt = _build_user_prompt(
        name=name,
        topics=topics or [],
        country=country,
        city=city,
        is_virtual=is_virtual,
    )
    req = ChatRequest(
        messages=[
            ChatMessage(role="system", content=get_settings().prompt_conference_enrichment),
            ChatMessage(role="user", content=user_prompt),
        ],
        purpose="enrich:conference",
        temperature=0.2,
        max_tokens=180,
    )
    try:
        resp = await get_llm_client().chat(req, db=db)
    except Exception as exc:
        log.warning(
            "enrichment.llm_failed",
            name=name[:60],
            error=str(exc)[:200],
        )
        return None
    text = (resp.content or "").strip()
    if not text:
        return None
    # Sanity bound: anything over 600 chars is the LLM rambling.
    if len(text) > 600:
        text = text[:600].rsplit(".", 1)[0] + "."
    return text


_YEAR = re.compile(r"\b(19|20)\d{2}\b")


_ORDINAL = re.compile(r"\b\d{1,3}(st|nd|rd|th)\b", re.I)


_EDITION = re.compile(r"\b(annual|edition|vol\.?|volume|no\.?|number)\b", re.I)


_EVENT_TYPE = re.compile(
    r"\b(conference|conf\.?|summit|symposium|workshop|meetup|expo|forum|"
    r"congress|days?|week)\b",
    re.I,
)


_REGION = re.compile(
    r"\b(europe|eu|emea|north america|na|namerica|latam|apac|asia[- ]?pacific|"
    r"asia|china|japan|korea|india|australia|anz|africa|middle east|"
    r"americas|us|usa|uk|global|international|world(?:wide)?)\b",
    re.I,
)


_PUNCT = re.compile(r"[^a-z0-9 ]+")


_SPACE = re.compile(r"\s+")


_STOP = {"the", "a", "an", "and", "of", "for", "on", "in", "at"}


_MIN_TOKENS_WITHOUT_TYPE = 2


def _drop_event_types(tokens: list[str]) -> list[str]:
    """Remove kind-of-gathering nouns, but only while the name survives it.

    "PyTorch Conference" -> "pytorch": still a name.
    "AI Summit"          -> "ai":      not a name, so "summit" stays.

    Keeping the noun on a short name is what stops "AI Summit" and "AI Expo"
    reading as the same event. It costs a false negative in the other
    direction — "AI Summit" and "AI Conference" will not link even if they
    are the same event — and that is the trade this module always makes: a
    missed link is recoverable, a wrong one silently applies one
    conference's history to another.
    """
    kept = [t for t in tokens if not _EVENT_TYPE.fullmatch(t)]
    return kept if len(kept) >= _MIN_TOKENS_WITHOUT_TYPE else tokens


def _base(name: str) -> str:
    """Lowercase, drop instance markers, squash punctuation, then tokenise."""
    s = (name or "").lower()
    s = _YEAR.sub(" ", s)
    s = _ORDINAL.sub(" ", s)
    s = _EDITION.sub(" ", s)
    # Punctuation becomes a space, not nothing: "KubeCon+CloudNativeCon"
    # must not fuse into one token, or it stops matching the spaced form.
    s = _PUNCT.sub(" ", s)
    s = _SPACE.sub(" ", s).strip()
    # Drop stop words and bare numbers. A standalone number is a volume or
    # an instance count ("Vol. 42"), never identity — but a number fused
    # into a token IS identity ("Conf42"), so only whole tokens go.
    tokens = [t for t in s.split() if t not in _STOP and not t.isdigit()]
    return " ".join(_drop_event_types(tokens))


def event_key(name: str) -> str:
    """Identity of the recurring event. Region KEPT, year dropped.

    ``KubeCon + CloudNativeCon Europe 2026`` -> ``cloudnativecon europe kubecon``

    Tokens are sorted so word order does not change identity — "CloudNativeCon
    + KubeCon EU" is the same event as "KubeCon + CloudNativeCon Europe".
    """
    return " ".join(sorted(set(_base(name).split())))


def series_key(name: str) -> str:
    """Identity of the franchise. Region AND year dropped.

    ``KubeCon + CloudNativeCon North America 2026`` -> ``cloudnativecon kubecon``
    """
    s = _REGION.sub(" ", _base(name))
    s = _SPACE.sub(" ", s).strip()
    return " ".join(sorted(set(s.split())))


def same_event(a: str, b: str) -> bool:
    """True when two names refer to the same recurring event.

    Different years of one event are the SAME event — that is what makes
    "we attended this last year" meaningful.
    """
    ka, kb = event_key(a), event_key(b)
    return bool(ka) and ka == kb


def same_series(a: str, b: str) -> bool:
    """True when two names belong to the same franchise.

    Weaker than :func:`same_event`. KubeCon EU and KubeCon NA are the same
    series but not the same event.
    """
    ka, kb = series_key(a), series_key(b)
    return bool(ka) and ka == kb


def relationship(a: str, b: str) -> str:
    """``"same_event"`` | ``"same_series"`` | ``"unrelated"``.

    The single call site most consumers want, so nobody has to remember
    which of the two predicates is the stronger one.
    """
    if same_event(a, b):
        return "same_event"
    if same_series(a, b):
        return "same_series"
    return "unrelated"


_REGION_CANON: dict[str, str] = {
    "eu": "europe", "emea": "europe", "europe": "europe", "european": "europe",
    "na": "northamerica", "namerica": "northamerica", "us": "northamerica",
    "usa": "northamerica", "america": "northamerica", "americas": "northamerica",
    "apac": "asiapacific", "asia": "asiapacific", "pacific": "asiapacific",
    "latam": "latinamerica",
}


_TYPE_CANON: dict[str, str] = {
    "conf": "conference", "conference": "conference",
    "symp": "symposium", "symposium": "symposium",
    "wksp": "workshop", "workshop": "workshop",
}


def duplicate_key(name: str) -> str:
    """A LOOSE identity used only to ASK "are these the same conference?".

    Never use this to merge anything automatically. ``event_key`` stays the
    strict answer, and it errs toward missing a link because a false merge
    applies one conference's attendance history and decisions to another —
    silently, and hard to unpick.

    This is the opposite trade, and it is safe precisely because nothing
    acts on it: it errs toward FINDING a link, and a human confirms or
    dismisses. Three normalisations event_key deliberately does not do:

        region words collapse   "EU" and "Europe" become one token, so
                                KubeCon EU 2026 and KubeCon +
                                CloudNativeCon Europe 2026 match. Regions
                                are still KEPT, so EU and NA stay apart —
                                they really are different events.
        type words collapse     "Conf" and "Conference" become one.
        apostrophes vanish      "World's" and "Worlds" become one, rather
                                than event_key's "world s" vs "worlds".
    """
    s = (name or "").lower().replace("'", "").replace("’", "")
    s = _YEAR.sub(" ", s)
    s = _ORDINAL.sub(" ", s)
    s = _EDITION.sub(" ", s)
    s = _PUNCT.sub(" ", s)
    s = _SPACE.sub(" ", s).strip()

    tokens = [t for t in s.split() if t not in _STOP and not t.isdigit()]
    tokens = [_REGION_CANON.get(t, _TYPE_CANON.get(t, t)) for t in tokens]
    tokens = _drop_event_types(tokens)
    return " ".join(sorted(set(tokens)))


_MIN_SHARED_TOKENS = 2


def looks_like_duplicate(a: str, b: str) -> bool:
    """True when two names are worth ASKING a human about.

    Matches on an exact normalised token set, OR on one name's tokens
    being a strict SUBSET of the other's. The subset case is what catches
    an abbreviated name against its full form:

        "KubeCon EU 2026"                       {europe, kubecon}
        "KubeCon + CloudNativeCon Europe 2026"  {cloudnativecon, europe, kubecon}

    ``event_key`` refuses that link on purpose — its docstring says an
    abbreviated name does not link to its full form, because a false
    merge is worse than a missed one. That reasoning holds for MERGING
    and inverts for FLAGGING: nothing here acts on the answer, so the
    cost of asking is one dismissal and the cost of not asking is a
    duplicate nobody notices.
    """
    ta, tb = set(duplicate_key(a).split()), set(duplicate_key(b).split())
    if not ta or not tb:
        return False
    if ta == tb:
        return True
    shared = ta & tb
    if len(shared) < _MIN_SHARED_TOKENS:
        return False
    return ta < tb or tb < ta


# ==========================================================================
# series.py
# ==========================================================================


async def create_series(
    db: AsyncSession,
    *,
    canonical_name: str,
    aliases: list[str] | None = None,
    description: str = "",
    typical_month: int | None = None,
    typical_topics: list[str] | None = None,
    homepage: str | None = None,
    actor_label: str = "system",
) -> ConferenceSeries:
    row = ConferenceSeries(
        canonical_name=canonical_name.strip(),
        aliases=list(aliases or []),
        description=description.strip(),
        typical_month=typical_month,
        typical_topics=list(typical_topics or []),
        homepage=homepage,
    )
    db.add(row)
    try:
        await db.flush()
    except IntegrityError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"A series named {canonical_name!r} already exists.",
        ) from exc
    await db.refresh(row)
    await write_audit(
        db,
        action="series.create",
        target_type="conference_series",
        target_id=row.id,
        before=None,
        after=model_to_audit_dict(row),
        actor_label=actor_label,
    )
    log.info("series.created", series_id=str(row.id), name=row.canonical_name)
    return row


async def update_series(
    db: AsyncSession,
    series_id: UUID,
    *,
    canonical_name: str | None = None,
    aliases: list[str] | None = None,
    description: str | None = None,
    typical_month: int | None = None,
    typical_topics: list[str] | None = None,
    homepage: str | None = None,
    is_active: bool | None = None,
    actor_label: str = "system",
) -> ConferenceSeries:
    row = await db.get(ConferenceSeries, series_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"No conference_series {series_id}")
    before = model_to_audit_dict(row)

    if canonical_name is not None:
        row.canonical_name = canonical_name.strip()
    if aliases is not None:
        row.aliases = list(aliases)
    if description is not None:
        row.description = description.strip()
    if typical_month is not None:
        row.typical_month = typical_month
    if typical_topics is not None:
        row.typical_topics = list(typical_topics)
    if homepage is not None:
        row.homepage = homepage
    if is_active is not None:
        row.is_active = is_active

    try:
        await db.flush()
    except IntegrityError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Series rename collides with an existing canonical_name.",
        ) from exc
    await db.refresh(row)
    await write_audit(
        db,
        action="series.update",
        target_type="conference_series",
        target_id=row.id,
        before=before,
        after=model_to_audit_dict(row),
        actor_label=actor_label,
    )
    return row


async def deactivate_series(
    db: AsyncSession, series_id: UUID, *, actor_label: str = "system"
) -> ConferenceSeries:
    row = await db.get(ConferenceSeries, series_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"No conference_series {series_id}")
    if not row.is_active:
        return row
    before = model_to_audit_dict(row)
    row.is_active = False
    await db.flush()
    await db.refresh(row)
    await write_audit(
        db,
        action="series.deactivate",
        target_type="conference_series",
        target_id=row.id,
        before=before,
        after=model_to_audit_dict(row),
        actor_label=actor_label,
    )
    return row


async def assign_conference_to_series(
    db: AsyncSession,
    series_id: UUID,
    conference_id: UUID,
    *,
    actor_label: str = "system",
) -> Conference:
    """Set ``conferences.series_id``; recompute the matcher for that
    conference asynchronously so the past-attendance bonus reflects the
    new link in the dashboard."""
    series = await db.get(ConferenceSeries, series_id)
    if series is None:
        raise HTTPException(status_code=404, detail=f"No conference_series {series_id}")
    if not series.is_active:
        raise HTTPException(
            status_code=409,
            detail=f"Series {series_id} is deactivated; reactivate before assigning.",
        )

    conf = await db.get(Conference, conference_id)
    if conf is None:
        raise HTTPException(status_code=404, detail=f"No conference {conference_id}")

    before = model_to_audit_dict(conf)
    conf.series_id = series.id
    await db.flush()
    await db.refresh(conf)
    await write_audit(
        db,
        action="series.assign",
        target_type="conference",
        target_id=conf.id,
        before=before,
        after=model_to_audit_dict(conf),
        actor_label=actor_label,
    )
    _enqueue_matcher_recompute(conf.id)
    return conf


async def unassign_conference_from_series(
    db: AsyncSession,
    conference_id: UUID,
    *,
    actor_label: str = "system",
) -> Conference:
    conf = await db.get(Conference, conference_id)
    if conf is None:
        raise HTTPException(status_code=404, detail=f"No conference {conference_id}")
    if conf.series_id is None:
        return conf
    before = model_to_audit_dict(conf)
    conf.series_id = None
    await db.flush()
    await db.refresh(conf)
    await write_audit(
        db,
        action="series.unassign",
        target_type="conference",
        target_id=conf.id,
        before=before,
        after=model_to_audit_dict(conf),
        actor_label=actor_label,
    )
    _enqueue_matcher_recompute(conf.id)
    return conf


@dataclass(slots=True)
class OrphanLinkResult:
    linked: int
    skipped: int

    def to_dict(self) -> dict:
        return {"linked": self.linked, "skipped": self.skipped}


async def link_conference_series_orphans(db: AsyncSession) -> OrphanLinkResult:
    """Attach conferences with no series to the series they belong to.

    Identity comes from ``series_identity.same_series`` — the same rule the
    matcher and the attendance history use, so a conference cannot be
    linked here and treated as unrelated there.

    No threshold, no needs_review bucket. Two names either share a series
    key or they do not; a confidence score between them was never something
    a person could act on. Caller commits.
    """

    orphans = (
        (await db.execute(select(Conference).where(Conference.series_id.is_(None))))
        .scalars()
        .all()
    )
    if not orphans:
        return OrphanLinkResult(linked=0, skipped=0)

    series_rows = (
        await db.execute(
            select(ConferenceSeries.id, ConferenceSeries.canonical_name).where(
                ConferenceSeries.is_active.is_(True)
            )
        )
    ).all()

    linked = skipped = 0
    for conf in orphans:
        match = next(
            (sid for sid, canonical in series_rows if same_series(conf.name, canonical)),
            None,
        )
        if match is None:
            skipped += 1
            continue
        conf.series_id = match
        await db.flush()
        log.info("conference.series.linked", conference_id=str(conf.id), name=conf.name)
        linked += 1

    return OrphanLinkResult(linked=linked, skipped=skipped)


def _enqueue_matcher_recompute(conference_id: UUID) -> None:
    """Local-import the scheduler + task to avoid a top-level import cycle
    (scheduler -> tasks -> series_service -> scheduler)."""

    enqueue_task(
        "run_fit_match",
        job_id=f"match-{conference_id}",
        kwargs={"conference_id": str(conference_id)},
    )


# ==========================================================================
# participation_service.py
# ==========================================================================


async def _require_conference(db: AsyncSession, conference_id: UUID) -> Conference:
    row = await db.get(Conference, conference_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Conference not found")
    return row


async def _resolve_person_label(
    db: AsyncSession, *, sme_id: UUID | None, person_label: str
) -> str:
    """Validate the SME and fall back to their name for a blank label."""
    if sme_id is None:
        return person_label
    sme = await db.get(Sme, sme_id)
    if sme is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "SME not found")
    return person_label or sme.full_name


def _to_read(row: Participation, *, today: date | None = None) -> ParticipationRead:
    """Serialise a row, computing ``has_attended`` rather than storing it.

    A person counts as having attended when someone confirmed it
    (``attended_at``), OR when their departure date has passed. Those are
    the two routes the operator described: "either by acknowledgement of
    the dates attending or by setting it to attended".

    Derived on read on purpose. Storing it would mean something has to
    write it, and the only thing that could write it on a schedule is a
    cron — which is exactly how the decay pass ended up overwriting
    recorded decisions. A predicate cannot drift.
    """
    ref = today or date.today()
    attended = row.attended_at is not None or (
        row.departs_on is not None and row.departs_on < ref
    )
    read = ParticipationRead.model_validate(row)
    return read.model_copy(update={"has_attended": attended})


async def mark_attended(
    db: AsyncSession, participation_id: UUID, *, attended: bool
) -> ParticipationRead:
    """Confirm (or un-confirm) that this person actually went.

    Un-confirming is supported because the alternative is that a
    mis-click is permanent, and a record nobody can correct stops being
    trusted. Clearing ``attended_at`` does not clear the dates, so a row
    whose departure has passed still reads as attended — the derived
    answer is the honest one.
    """
    row = await db.get(Participation, participation_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Participation not found")
    row.attended_at = datetime.now(tz=UTC) if attended else None
    await db.flush()
    await db.refresh(row)
    log.info(
        "participation.attendance_marked",
        participation_id=str(participation_id),
        attended=attended,
    )
    return _to_read(row)


async def list_participation(
    db: AsyncSession, conference_id: UUID
) -> list[ParticipationRead]:
    await _require_conference(db, conference_id)
    rows = (
        (
            await db.execute(
                select(Participation)
                .where(Participation.conference_id == conference_id)
                .order_by(Participation.activity, Participation.person_label)
            )
        )
        .scalars()
        .all()
    )
    return [_to_read(r) for r in rows]


async def add_participation(
    db: AsyncSession, conference_id: UUID, payload: ParticipationCreate
) -> ParticipationRead:
    await _require_conference(db, conference_id)
    label = await _resolve_person_label(
        db, sme_id=payload.sme_id, person_label=payload.person_label
    )

    row = Participation(
        conference_id=conference_id,
        sme_id=payload.sme_id,
        person_label=label,
        activity=payload.activity,
        talk_id=payload.talk_id,
        arrives_on=payload.arrives_on,
        departs_on=payload.departs_on,
        notes=payload.notes or "",
    )
    db.add(row)
    await db.flush()
    # created_at/updated_at come from server defaults, which flush does not
    # load back. Without this, reading them raises MissingGreenlet.
    await db.refresh(row)
    log.info(
        "participation.added",
        conference_id=str(conference_id),
        activity=payload.activity,
        person=label,
    )
    return _to_read(row)


async def update_participation(
    db: AsyncSession, participation_id: UUID, payload: ParticipationUpdate
) -> ParticipationRead:
    row = await db.get(Participation, participation_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Participation not found")

    row.person_label = await _resolve_person_label(
        db, sme_id=payload.sme_id, person_label=payload.person_label
    )
    row.sme_id = payload.sme_id
    row.activity = payload.activity
    row.talk_id = payload.talk_id
    row.arrives_on = payload.arrives_on
    row.departs_on = payload.departs_on
    row.notes = payload.notes or ""
    await db.flush()
    await db.refresh(row)
    return _to_read(row)


async def delete_participation(db: AsyncSession, participation_id: UUID) -> None:
    row = await db.get(Participation, participation_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Participation not found")
    await db.delete(row)
    await db.flush()


async def get_attendance(db: AsyncSession, conference_id: UUID) -> AttendanceSummary:
    row = await _require_conference(db, conference_id)
    return AttendanceSummary(
        edition_year=row.edition_year,
        spend_usd=row.spend_usd,
        leads_generated=row.leads_generated,
        audience_size_estimate=row.audience_size_estimate,
        attendance_verdict=row.attendance_verdict,
        attendance_notes=row.attendance_notes,
    )


async def set_attendance(
    db: AsyncSession, conference_id: UUID, payload: AttendanceSummary
) -> AttendanceSummary:
    row = await _require_conference(db, conference_id)
    row.edition_year = payload.edition_year
    row.spend_usd = payload.spend_usd
    row.leads_generated = payload.leads_generated
    row.audience_size_estimate = payload.audience_size_estimate
    row.attendance_verdict = payload.attendance_verdict
    row.attendance_notes = payload.attendance_notes
    await db.flush()
    log.info("participation.attendance_updated", conference_id=str(conference_id))
    return payload


# ==========================================================================
# import_past.py — Excel/CSV import of already-attended conferences
# ==========================================================================
#
# Past attendance is matcher food: it feeds the past-attendance dimension,
# the series-memory boosts, the "been before" filters, and the analytics
# outcome charts. This importer takes one spreadsheet row per conference,
# creates (or finds) the conference, records who went as attended
# participation rows — linked to SME records when the name matches the
# roster — and stores the outcome fields.
#
# IMPORT_COLUMNS is the single source of truth for the format: the popup
# in the UI renders it, the downloadable template is generated from it,
# and the parser reads by it. Change it here and all three move together.

IMPORT_COLUMNS: Final[list[dict]] = [
    {"key": "name", "label": "Conference name", "required": True, "example": "KubeCon NA 2025"},
    {"key": "start_date", "label": "Start date (YYYY-MM-DD)", "required": False, "example": "2025-11-12"},
    {"key": "end_date", "label": "End date (YYYY-MM-DD)", "required": False, "example": "2025-11-15"},
    {"key": "city", "label": "City", "required": False, "example": "Atlanta"},
    {"key": "country", "label": "Country (2-letter code)", "required": False, "example": "US"},
    {"key": "website", "label": "Website", "required": False, "example": "https://kubecon.io"},
    {
        "key": "attended_by",
        "label": "Who attended (names separated by ;)",
        "required": False,
        "example": "Isaac Tigges; Jane Doe",
    },
    {"key": "spend_usd", "label": "Actual cost in USD (number)", "required": False, "example": "4500"},
    {"key": "leads_generated", "label": "Leads generated (number)", "required": False, "example": "32"},
    {
        "key": "worth_it",
        "label": "Worth it? (yes / no / unsure)",
        "required": False,
        "example": "yes",
    },
    {"key": "notes", "label": "Notes", "required": False, "example": "Great hallway track"},
]

_WORTH_MAP: Final[dict[str, str]] = {
    "yes": "would_attend",
    "y": "would_attend",
    "would_attend": "would_attend",
    "no": "would_not_attend",
    "n": "would_not_attend",
    "would_not_attend": "would_not_attend",
    "unsure": "unsure",
    "maybe": "unsure",
}


def build_import_template() -> bytes:
    """.xlsx with the header row plus one example row, from IMPORT_COLUMNS."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Past conferences"
    for col, spec in enumerate(IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=spec["key"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="EE0000")
        ws.cell(row=2, column=col, value=spec["example"])
        ws.column_dimensions[get_column_letter(col)].width = max(
            len(str(spec["key"])), len(str(spec["example"]))
        ) + 2
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _parse_import_rows(filename: str, raw: bytes) -> list[dict[str, str]]:
    """File bytes -> list of {column_key: cell_string}. Header-driven, so
    column order in the sheet does not matter and extra columns are ignored."""
    import csv
    import io

    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = [dict(r) for r in reader]
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        raw_rows = [
            {header[i]: cell for i, cell in enumerate(row) if i < len(header)}
            for row in rows_iter
        ]
    else:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported file type — upload .xlsx or .csv (download the template).",
        )

    known = {spec["key"] for spec in IMPORT_COLUMNS}
    out: list[dict[str, str]] = []
    for r in raw_rows:
        cleaned: dict[str, str] = {}
        for k, v in r.items():
            key = str(k or "").strip().lower().replace(" ", "_")
            if key not in known or v is None:
                continue
            if isinstance(v, datetime | date):
                cleaned[key] = v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
            else:
                cleaned[key] = str(v).strip()
        if any(cleaned.values()):
            out.append(cleaned)
    return out


async def import_past_conferences(
    db: AsyncSession, *, filename: str, raw: bytes, actor_label: str = "import"
) -> dict:
    """One spreadsheet -> conferences + attended participation + outcomes.

    Idempotent by slug: a row whose conference already exists attaches its
    attendance/outcome data to the existing record instead of duplicating
    it, so re-uploading a corrected sheet is safe.
    """
    from app.services.embeddings import embed_owner
    from app.services.extraction import build_slug, year_for

    rows = _parse_import_rows(filename, raw)
    if len(rows) > 500:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Too many rows ({len(rows)}); the importer caps at 500 per file.",
        )

    smes = (
        await db.execute(select(Sme).where(Sme.is_active.is_(True)))
    ).scalars().all()
    sme_by_name = {s.full_name.strip().lower(): s for s in smes}

    def _parse_date(v: str | None) -> date | None:
        if not v:
            return None
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None

    def _parse_int(v: str | None) -> int | None:
        if not v:
            return None
        try:
            return int(float(v))
        except ValueError:
            return None

    results: list[dict] = []
    created = updated = errors = 0
    for idx, r in enumerate(rows, start=2):  # row 1 is the header
        name = (r.get("name") or "").strip()
        if not name:
            errors += 1
            results.append({"row": idx, "name": "", "outcome": "error", "detail": "name is required"})
            continue

        start = _parse_date(r.get("start_date"))
        slug = build_slug(name, year_for(start))
        conf = (
            await db.execute(select(Conference).where(Conference.slug == slug))
        ).scalar_one_or_none()
        is_new = conf is None
        if is_new:
            conf = Conference(
                name=name,
                slug=slug,
                event_kind="corporate",
                start_date=start,
                end_date=_parse_date(r.get("end_date")),
                location_city=(r.get("city") or None),
                location_country=((r.get("country") or "").upper()[:2] or None),
                is_virtual=False,
                website=(r.get("website") or None),
                cfp_deadlines=[],
                cfp_topics_of_interest=[],
                topics=[],
                confidence_score=1.0,
                # A conference somebody already went to needs no review
                # funnel — it is history, not a candidate.
                status="approved",
            )
            db.add(conf)
            await db.flush()

        # Outcome fields — fill only what the sheet provides; never wipe
        # existing values with blanks.
        spend = _parse_int(r.get("spend_usd"))
        leads = _parse_int(r.get("leads_generated"))
        verdict = _WORTH_MAP.get((r.get("worth_it") or "").strip().lower())
        notes = (r.get("notes") or "").strip()
        if spend is not None:
            conf.spend_usd = spend
        if leads is not None:
            conf.leads_generated = leads
        if verdict:
            conf.attendance_verdict = verdict
        if notes:
            conf.attendance_notes = notes

        # Attended participation rows, SME-linked when the name matches.
        existing_labels = {
            label.strip().lower()
            for (label,) in (
                await db.execute(
                    select(Participation.person_label).where(
                        Participation.conference_id == conf.id
                    )
                )
            ).all()
        }
        attendees = [
            a.strip() for a in (r.get("attended_by") or "").split(";") if a.strip()
        ]
        linked = 0
        for person in attendees:
            if person.lower() in existing_labels:
                continue
            sme = sme_by_name.get(person.lower())
            db.add(
                Participation(
                    conference_id=conf.id,
                    sme_id=sme.id if sme else None,
                    person_label=person,
                    activity="attend",
                    attended_at=datetime.now(tz=UTC),
                )
            )
            if sme:
                linked += 1
        await db.flush()

        if is_new:
            created += 1
        else:
            updated += 1
        results.append(
            {
                "row": idx,
                "name": name,
                "outcome": "created" if is_new else "updated existing",
                "detail": (
                    f"{len(attendees)} attendee(s), {linked} matched to SME records"
                    if attendees
                    else "no attendees listed"
                ),
            }
        )

        # Best-effort embed so the imported history joins the matcher
        # corpus; a failed embed must not fail the import.
        if is_new:
            try:
                await embed_owner(
                    db,
                    owner_type="conference",
                    owner_id=conf.id,
                    text=conference_embed_text(conf),
                    purpose="embed:conference_import",
                )
            except Exception as exc:
                log.warning(
                    "conference.import.embed_failed",
                    slug=slug,
                    error=f"{type(exc).__name__}: {exc}",
                )

    log.info(
        "conference.import_past",
        actor=actor_label,
        created=created,
        updated=updated,
        errors=errors,
    )
    await db.commit()
    return {
        "total_rows": len(rows),
        "created": created,
        "updated_existing": updated,
        "errors": errors,
        "results": results,
    }
