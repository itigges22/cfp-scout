"""CSV linter for the AI BU Developer Marketing events spreadsheet.

Mirrors the parsing logic in a teammate's calendar-sync utility's
``main.py`` (the ``clean_event_data`` + ``parse_date`` functions) so a CSV
that runs cleanly through the calendar-sync cron also runs cleanly through
Scout's importer. Any change to that upstream parser should be mirrored
here.

Output is structured `LintedEvent` dicts; the caller (mapper.py) translates
those into Scout's `past_conferences` and `conferences` rows.

When the input CSV doesn't have the expected column shape we raise
`LinterFormatError` so the caller can decide whether to fall back to a
generic Docling parse + LLM extract.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

# Exact column names from the upstream spreadsheet. Renaming any of these
# upstream would break the cron — so we mirror them verbatim.
REQUIRED_COLUMNS = {
    "Event Name",
    "Complete",
    "Type",
    "Start Date",
    "End Date",
    "City",
    "Country",
    "AI BU On-Site Staff",
    "Description",
    "Activities",
}

# Quarter-marker rows the upstream parser skips outright.
QUARTER_MARKERS = {"Q1", "Q2", "Q3", "Q4"}

_MONTHS_ONLY = {
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
}


class LinterFormatError(RuntimeError):
    """Raised when the CSV doesn't have the expected calendar-sync columns.
    The endpoint catches this and falls back to Docling + LLM extraction."""


@dataclass(slots=True)
class LintedEvent:
    complete: bool
    type: str
    name: str
    start_date: datetime | None
    end_date: datetime | None
    city: str
    country: str
    attendees_raw: str  # comma-separated names; mapper resolves to SME UUIDs
    description: str
    activities: str
    source_row: int  # 2-based (header is row 1) so warnings line up with the file
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class LintResult:
    events: list[LintedEvent]
    skipped: int  # rows the linter dropped (empty / quarter marker / no parsable date)
    warnings: list[str]  # file-level warnings (vs per-row in LintedEvent.warnings)


def parse_date(value: str | None, *, fallback_year: int) -> datetime | None:
    """Port of `parse_date` from main.py.

    Returns None for blanks, 'TBD', 'week of …', or bare month names. Strips
    English ordinal suffixes (st/nd/rd/th) before strptime'ing.
    """
    if not value:
        return None
    s = value.strip()
    if not s or s.upper() == "TBD":
        return None
    if s.lower() in _MONTHS_ONLY:
        return None
    if "TBD" in s or "week" in s.lower():
        return None
    try:
        clean = re.sub(r"(st|nd|rd|th)\s*$", "", s)
        return datetime.strptime(f"{clean} {fallback_year}", "%B %d %Y")
    except ValueError:
        pass
    # Try ISO-ish too (the upstream parser doesn't, but operators sometimes
    # paste YYYY-MM-DD; cheap to support without breaking the original behavior).
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def lint_calendar_sync_csv(
    content: bytes, *, fallback_year: int = 2026
) -> LintResult:
    """Lint a CSV exported from the AI BU Developer Marketing spreadsheet.

    Args:
        content: raw CSV bytes (UTF-8 expected; we don't bend over backwards
                 on encoding because the upstream is always Google Sheets).
        fallback_year: year to assume when the date column has no year
                       (e.g. "June 14th"). Defaults to 2026 to match the
                       current spreadsheet's tab. Should be set per upload
                       by the endpoint if it ever needs to be different.

    Returns: LintResult.

    Raises: LinterFormatError if the CSV is missing required columns or
            isn't parseable as CSV at all — caller should then try the
            Docling fallback.
    """
    # XLSX shortcut: the upstream spreadsheet's natural format is XLSX,
    # not CSV. If the upload is an XLSX, convert it to a CSV string in
    # memory via openpyxl and feed THAT to the same parser — orders of
    # magnitude faster than routing through Docling + the LLM extractor,
    # and preserves the strict-column-shape guarantees.
    if content.startswith(b"PK\x03\x04"):
        try:
            text = _xlsx_to_csv_text(content)
        except Exception as exc:  # noqa: BLE001 — surface as a format error
            raise LinterFormatError(
                f"file looks like an XLSX but openpyxl couldn't read it: {exc}"
            ) from exc
    else:
        # utf-8-sig strips a BOM if Excel/Numbers added one. Fall back to
        # replace-decoded errors so a single funny byte in a description
        # cell doesn't crash the whole import.
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("utf-8", errors="replace")

    # Normalize line endings BEFORE handing to csv. Google Sheets exports
    # as \n, but files round-tripped through Excel / Outlook / Word can
    # arrive with \r\n or bare \r — and the csv module raises "new-line
    # character seen in unquoted field" the moment it sees a stray \r
    # inside what it thinks is an unquoted cell. Collapsing everything to
    # \n means csv only ever has to recognize one line terminator, and
    # any real embedded-newline-inside-quoted-field still works because
    # csv handles \n inside quotes natively.
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # newline="" on StringIO tells it not to do its own translation, so
    # csv sees the bytes verbatim. Standard pattern from the csv docs.
    reader = csv.DictReader(io.StringIO(text, newline=""))
    if reader.fieldnames is None:
        raise LinterFormatError("CSV has no header row")

    header_set = {h.strip() for h in reader.fieldnames if h}
    missing = REQUIRED_COLUMNS - header_set
    if missing:
        raise LinterFormatError(
            f"CSV missing required columns: {sorted(missing)}. "
            f"Got: {sorted(header_set)}"
        )

    events: list[LintedEvent] = []
    skipped = 0
    file_warnings: list[str] = []

    # csv reader starts at the line AFTER the header, so the first row's
    # 1-indexed file line is 2.
    for row_idx, row in enumerate(reader, start=2):
        event_name = (row.get("Event Name") or "").strip()
        if not event_name or event_name in QUARTER_MARKERS:
            skipped += 1
            continue

        warnings: list[str] = []
        complete = (row.get("Complete") or "").strip().upper() == "TRUE"
        event_type = (row.get("Type") or "").strip()
        start_raw = (row.get("Start Date") or "").strip()
        end_raw = (row.get("End Date") or "").strip()
        city = (row.get("City") or "").strip()
        country = (row.get("Country") or "").strip()
        attendees = (row.get("AI BU On-Site Staff") or "").strip()
        description = (row.get("Description") or "").strip()
        activities = (row.get("Activities") or "").strip()

        start_date = parse_date(start_raw, fallback_year=fallback_year)
        end_date = parse_date(end_raw, fallback_year=fallback_year)

        if start_date is None:
            # Upstream behavior: print warning + skip the row. We surface as
            # a per-row warning attached to a dropped event — but since we
            # have no start date, the dropped row's metadata is mostly
            # useless. Just bump the skip counter and add a file warning.
            file_warnings.append(
                f"row {row_idx} '{event_name}': unparseable start date "
                f"{start_raw!r} — row dropped"
            )
            skipped += 1
            continue

        if end_date is None:
            warnings.append("end date missing — using start date (single-day event)")
            end_date = start_date

        events.append(
            LintedEvent(
                complete=complete,
                type=event_type,
                name=event_name,
                start_date=start_date,
                end_date=end_date,
                city=city,
                country=country,
                attendees_raw=attendees,
                description=description,
                activities=activities,
                source_row=row_idx,
                warnings=warnings,
            )
        )

    return LintResult(events=events, skipped=skipped, warnings=file_warnings)


# ---------------------------------------------------------------------------
# Helpers shared with the mapper
# ---------------------------------------------------------------------------
def _xlsx_to_csv_text(content: bytes) -> str:
    """Convert XLSX bytes to a CSV string.

    Picks the first sheet that has the required columns; falls back to the
    active sheet if none match. Cells are stringified, dates are formatted
    as the upstream spreadsheet's "Month Day" format ("June 14") so the
    same `parse_date` rules apply (we then tack on the year in the linter).

    Returns the CSV body, ready to hand to csv.DictReader.
    """
    from datetime import date as _date, datetime as _dt
    from io import BytesIO, StringIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)

    # Prefer the first sheet whose headers contain our required columns.
    target = None
    for ws_candidate in wb.worksheets:
        # Read the header row only (first non-empty row) for the column check.
        rows_iter = ws_candidate.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        header_set = {str(h).strip() for h in header if h is not None}
        if REQUIRED_COLUMNS.issubset(header_set):
            target = ws_candidate
            break
    if target is None:
        target = wb.active

    out = StringIO()
    writer = csv.writer(out, lineterminator="\n")
    for row in target.iter_rows(values_only=True):
        rendered: list[str] = []
        for cell in row:
            if cell is None:
                rendered.append("")
            elif isinstance(cell, _dt):
                # Match the upstream "June 14th" / "June 14" shape so the
                # existing parse_date branch picks it up. Ordinal suffix
                # isn't required — strptime tolerates both.
                rendered.append(cell.strftime("%B %d"))
            elif isinstance(cell, _date):
                rendered.append(cell.strftime("%B %d"))
            elif isinstance(cell, bool):
                # Excel booleans → upstream TRUE/FALSE strings.
                rendered.append("TRUE" if cell else "FALSE")
            else:
                rendered.append(str(cell))
        writer.writerow(rendered)
    return out.getvalue()


def split_attendees(raw: str) -> list[str]:
    """Comma-or-semicolon split, trimmed, blanks dropped.

    The upstream sheet uses commas in `AI BU On-Site Staff`; some operators
    sometimes paste in semicolons. Tolerate both.
    """
    if not raw:
        return []
    pieces: Iterable[str] = re.split(r"[;,]", raw)
    return [p.strip() for p in pieces if p.strip()]
