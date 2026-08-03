"""What the operator reads: the CFP digest and the conference brief.

WHAT THIS DOES
    Two generated documents. The digest is the recurring "these deadlines
    are coming up" summary that drives notifications. The brief is the
    long-form per-conference write-up — why it scored, who should go,
    what to pitch.

HOW IT CONNECTS
    Called by   api/v1/briefs.py, api/v1/notifications.py,
                tasks.py
    Reads       conferences, matches, smes, talks, pillars
    Helpers     services/llm.py, services/matcher/, services/embeddings.py

WORTH KNOWING
    Both answer "what should a human do about this?" from the same rows
    at different lengths, and both break the same way when the matcher
    changes shape.
"""

from __future__ import annotations

import asyncio
import time
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime, timedelta
from typing import Any
from uuid import UUID

import structlog
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import (
    Conference,
    ConferencePillar,
    ConferenceSeries,
    ConferenceSource,
    Decision,
    Match,
    MessagingDocument,
    Notification,
    Participation,
    Sme,
    StrategicPillar,
    Talk,
    TalkSubmission,
)
from app.services import conferences as cs
from app.services.embeddings import similar_chunks
from app.services.matcher import ALGORITHM_VERSION, run_fit_match
from app.settings import get_settings

log = structlog.get_logger("scout.reports")


# ==========================================================================
# digest.py
# ==========================================================================


# Today and tomorrow only. The digest is a deadline ALARM, not a browsing
# surface — the finder already has a "CFP closes within N days" filter for
# planning ahead. An empty digest means nothing is urgent, which is a
# valid and common answer; padding it with 30 days of lead time buried
# the one deadline that actually needed action today.
BUCKET_BOUNDS = [(0, 0, "today"), (1, 1, "tomorrow")]




@dataclass(slots=True, frozen=True)
class DigestEntry:
    """One (conference, deadline) row inside a bucket."""

    conference_id: str
    name: str
    slug: str
    status: str
    overall_score: float | None
    deadline_kind: str  # "submission" / "abstract" / "workshop" / ...
    deadline_date: str  # ISO
    days_until: int
    top_sme_id: str | None
    top_sme_name: str | None
    website: str | None
    location: str | None  # "city, country" or "virtual" or None


@dataclass(slots=True)
class DigestResult:
    generated_at: str
    notification_id: str | None
    buckets: dict[str, list[DigestEntry]] = field(default_factory=dict)
    stats: dict[str, int] = field(default_factory=dict)

    def to_payload(self) -> dict:
        """Shape persisted into ``notifications.payload``."""
        return {
            "generated_at": self.generated_at,
            "buckets": {k: [asdict(e) for e in v] for k, v in self.buckets.items()},
            "stats": self.stats,
        }

    def to_stats(self) -> dict:
        return {
            "generated_at": self.generated_at,
            "notification_id": self.notification_id,
            "n_entries_total": sum(self.stats.values()),
            "by_bucket": self.stats,
        }


async def build_cfp_digest(db: AsyncSession, *, today: date | None = None) -> DigestResult:
    """Walk the conference + cfp_deadlines space, bucket, persist, return.

    Caller commits. Marks any prior un-seen `cfp_digest` notifications as
    `seen=true` before inserting the new one — bell stays at 1.
    """
    today = today or date.today()
    horizon = today + timedelta(days=1)

    # Pull conferences + matches in one join. We only need rows whose
    # status is eligible AND whose JSONB cfp_deadlines array is non-empty.
    # The JSONB check via ``func.jsonb_array_length(... ) > 0`` is
    # cheap with an expression index, but for phase-1 volume we can just
    # filter in Python — much simpler code, same result.
    rows = (
        await db.execute(
            select(Conference, Match)
            .outerjoin(
                Match,
                (Match.conference_id == Conference.id)
                & (Match.algorithm_version == ALGORITHM_VERSION),
            )
            .where(Conference.status.in_(list(cs.DIGEST_ELIGIBLE)))
        )
    ).all()

    # Pre-fetch SME names for the "top SME" hint. We map top_sme_id =
    # the first recommended_sme_id from the match row (matcher already
    # sorted them by composite score).
    sme_ids: set[UUID] = set()
    for _, m in rows:
        if m and m.recommended_sme_ids:
            sme_ids.add(m.recommended_sme_ids[0])
    sme_name_by_id: dict[UUID, str] = {}
    if sme_ids:
        sme_rows = (
            await db.execute(select(Sme.id, Sme.full_name).where(Sme.id.in_(list(sme_ids))))
        ).all()
        sme_name_by_id = {sid: name for sid, name in sme_rows}

    # Build entries: explode (conf, deadlines_array) into per-deadline rows.
    entries: list[DigestEntry] = []
    for conf, match in rows:
        deadlines = list(conf.cfp_deadlines or [])
        # cfp_deadlines is the rich multi-deadline array, but discovery
        # rarely fills it — cfp_close_at is the canonical field the finder,
        # filters and exports all use. Reading only the array meant the
        # digest saw 2 of the 67 conferences actually closing this month
        # and shipped empty forever.
        if not deadlines and conf.cfp_close_at is not None:
            deadlines = [{"kind": "cfp", "deadline_date": conf.cfp_close_at.isoformat()}]
        if not deadlines:
            continue
        top_sme_id = match.recommended_sme_ids[0] if (match and match.recommended_sme_ids) else None
        top_sme_name = sme_name_by_id.get(top_sme_id) if top_sme_id is not None else None
        for d in deadlines:
            iso = d.get("deadline_date")
            if not iso:
                continue
            try:
                dd = date.fromisoformat(iso)
            except (TypeError, ValueError):
                continue
            if not (today <= dd <= horizon):
                continue
            entries.append(
                DigestEntry(
                    conference_id=str(conf.id),
                    name=conf.name,
                    slug=conf.slug,
                    status=conf.status,
                    overall_score=(float(match.overall_score) if match else None),
                    deadline_kind=d.get("kind") or "other",
                    deadline_date=iso,
                    days_until=(dd - today).days,
                    top_sme_id=str(top_sme_id) if top_sme_id else None,
                    top_sme_name=top_sme_name,
                    website=conf.website,
                    location=_pretty_location(conf),
                )
            )

    # Bucket + rank.
    buckets: dict[str, list[DigestEntry]] = {key: [] for _, _, key in BUCKET_BOUNDS}
    for e in entries:
        for lo, hi, key in BUCKET_BOUNDS:
            if lo <= e.days_until <= hi:
                buckets[key].append(e)
                break
    for key in buckets:
        buckets[key].sort(
            key=lambda e: (
                -(e.overall_score or 0.0),  # higher score first
                e.deadline_date,  # then earlier deadline first
            )
        )
        buckets[key] = buckets[key][:get_settings().digest_max_per_bucket]

    stats = {key: len(buckets[key]) for key in buckets}
    generated_at = datetime.now(tz=UTC).isoformat()

    # Mark prior un-seen digests as seen so the bell doesn't accumulate.
    await db.execute(
        update(Notification)
        .where(Notification.kind == "cfp_digest")
        .where(Notification.seen.is_(False))
        .values(seen=True)
    )

    result = DigestResult(
        generated_at=generated_at,
        notification_id=None,
        buckets=buckets,
        stats=stats,
    )
    # Persist as a fresh unread notification only if there's something to
    # surface; an empty digest doesn't need a bell badge.
    total = sum(stats.values())
    if total > 0:
        row = Notification(kind="cfp_digest", payload=result.to_payload(), seen=False)
        db.add(row)
        await db.flush()
        await db.refresh(row)
        result.notification_id = str(row.id)

    log.info(
        "digest.cfp.built",
        total_entries=total,
        bucket_today=stats.get("today", 0),
        bucket_tomorrow=stats.get("tomorrow", 0),
        notification_id=result.notification_id,
    )
    return result


def _pretty_location(c: Conference) -> str | None:
    if c.is_virtual:
        return "virtual"
    parts = [p for p in (c.location_city, c.location_country) if p]
    return ", ".join(parts) if parts else None


def to_markdown(result: DigestResult, *, today: date | None = None) -> str:
    """Format the digest as Markdown for the UI's copy-to-clipboard button.

    Pure function (no DB), so the frontend can also call this shape from a
    persisted notification.payload via the API if it wants server-rendered
    copy.
    """
    today = today or date.today()
    out: list[str] = [f"# Scout CFP Digest — {today.isoformat()}", ""]
    titles = {
        "today": "Closing TODAY",
        "tomorrow": "Closing tomorrow",
    }
    any_content = False
    for _, _, key in BUCKET_BOUNDS:
        entries = result.buckets.get(key, [])
        if not entries:
            continue
        any_content = True
        out.append(f"## {titles[key]}")
        out.append("")
        for e in entries:
            score = (
                f" (score {round((e.overall_score or 0) * 100)})"
                if e.overall_score is not None
                else ""
            )
            sme = f"; suggested SME: {e.top_sme_name}" if e.top_sme_name else ""
            kind_label = e.deadline_kind.replace("_", " ").title()
            out.append(f"- **{e.name}**{score} — {kind_label} closes {e.deadline_date}{sme}")
        out.append("")
    if not any_content:
        out.append("_No CFPs closing in the next 30 days._")
    return "\n".join(out).rstrip() + "\n"


# ==========================================================================
# brief.py
# ==========================================================================


CACHE_TTL_SECONDS = 300.0  # 5 minutes










class BriefNotFoundError(LookupError):
    """Conference does not exist."""


@dataclass(slots=True)
class _Entry:
    payload: dict[str, Any]
    built_at: float


_cache: dict[str, _Entry] = {}


_lock = asyncio.Lock()


def invalidate_cache(conference_id: UUID | str | None = None) -> None:
    """Drop cached briefs. If ``conference_id`` is given, drops only that
    conference's entry; otherwise clears everything."""
    if conference_id is None:
        _cache.clear()
        return
    _cache.pop(str(conference_id), None)


async def build_brief(
    db: AsyncSession,
    conference_id: UUID,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Build (and cache) the brief payload for ``conference_id``.

    Raises BriefNotFoundError if the conference doesn't exist.
    """
    key = str(conference_id)
    now = time.monotonic()
    cached = _cache.get(key)
    if not force and cached is not None and (now - cached.built_at) < CACHE_TTL_SECONDS:
        return cached.payload

    async with _lock:
        cached = _cache.get(key)
        if (
            not force
            and cached is not None
            and (time.monotonic() - cached.built_at) < CACHE_TTL_SECONDS
        ):
            return cached.payload

        payload = await _build(db, conference_id)
        _cache[key] = _Entry(payload=payload, built_at=time.monotonic())
        return payload


async def _build(db: AsyncSession, conference_id: UUID) -> dict[str, Any]:
    conference = await db.get(Conference, conference_id)
    if conference is None:
        raise BriefNotFoundError(f"No conference {conference_id}")

    match = (
        await db.execute(
            select(Match)
            .where(Match.conference_id == conference.id)
            .where(Match.algorithm_version == ALGORITHM_VERSION)
        )
    ).scalar_one_or_none()

    # Auto-run the matcher inline if no match exists yet for this algorithm
    # version. The brief endpoint should be self-sufficient — the user
    # shouldn't have to know about a separate "run matcher" step. The UI
    # already shows a loading spinner while this endpoint runs, so paying
    # the matcher cost here gives the user a single "open brief →
    # populated brief" interaction instead of an empty brief + manual
    # admin command.
    if match is None:
        try:

            log.info("brief.auto_match", conference_id=str(conference.id))
            await run_fit_match(db, conference.id)
            await db.commit()
            match = (
                await db.execute(
                    select(Match)
                    .where(Match.conference_id == conference.id)
                    .where(Match.algorithm_version == ALGORITHM_VERSION)
                )
            ).scalar_one_or_none()
        except Exception as exc:
            log.warning(
                "brief.auto_match_failed",
                conference_id=str(conference.id),
                error=str(exc)[:200],
            )

    series = None
    if conference.series_id is not None:
        series = await db.get(ConferenceSeries, conference.series_id)

    past = await _past_editions(db, conference)
    pillar = await _matched_pillar(db, conference.id)
    topics = await _top_topics(db, conference.id)
    attendees = await _attendees(db, conference_id)
    decision = await _latest_decision(db, conference.id)
    talking_docs = await _talking_points(db, conference)
    sources_count = await _sources_count(db, conference.id)

    settings = get_settings()
    now = datetime.now(tz=UTC)

    return {
        "conference_id": str(conference.id),
        "generated_at": now.isoformat(timespec="seconds"),
        "algorithm_version": ALGORITHM_VERSION,
        "scout_version": settings.scout_version if hasattr(settings, "scout_version") else "0.1.0",
        # --- 1. Header -------------------------------------------------
        "header": {
            "name": conference.name,
            "slug": conference.slug,
            "start_date": _iso_date(conference.start_date),
            "end_date": _iso_date(conference.end_date),
            "location_city": conference.location_city,
            "location_country": conference.location_country,
            "is_virtual": conference.is_virtual,
            "venue": conference.venue,
            "website": conference.website,
        },
        # --- 2. At a glance --------------------------------------------
        "at_a_glance": {
            "overall_score": _round(match.overall_score) if match else None,
            "fit_score": _round(match.fit_score) if match else None,
            "speaker_score": _round(match.speaker_score) if match else None,
            "overall_bucket": _bucket(match.overall_score) if match else None,
            "status": conference.status,
            "acceptance_rate_percent": conference.acceptance_rate_percent,
            "estimated_cost_usd": conference.estimated_cost_usd,
            "series": _series_summary(series, past) if series else None,
        },
        # --- 3. Why we're going ---------------------------------------
        "why": {
            "rationale_text": match.rationale_text if match else "",
            "matched_pillar": pillar,
            "top_topics": topics[:3],
        },
        # --- 4. Recommended attendee(s) -------------------------------
        "attendees": attendees,
        # --- 5. CFP info ----------------------------------------------
        "cfp": _cfp_section(conference),
        # --- 6. Past engagement ---------------------------------------
        "past_engagement": past,
        # --- 7. Talking points ----------------------------------------
        "talking_points": talking_docs,
        # --- 8. Logistics placeholder ---------------------------------
        # Real, shared, persisted. These used to be a localStorage key the
        # backend computed for the frontend, so the values lived in one
        # person's browser and vanished on a cache clear.
        "logistics": {
            "travel": conference.logistics_travel,
            "lodging": conference.logistics_lodging,
            "booth": conference.logistics_booth,
            "sponsorship": conference.logistics_sponsorship,
        },
        # --- 9. Footer -------------------------------------------------
        "footer": {
            "detail_url_path": f"/conferences/{conference.id}",
            "decision": _decision_summary(decision),
            "sources_count": sources_count,
        },
    }


async def _past_editions(db: AsyncSession, conference: Conference) -> list[dict]:
    """Earlier editions of this series that the team actually went to.

    Attendance is the participation rows, so an edition only appears here
    once somebody is recorded against it — and each person arrives with
    what they did, not just that they were present.
    """
    if conference.series_id is None:
        return []
    rows = (
        (
            await db.execute(
                select(Conference)
                .where(Conference.series_id == conference.series_id)
                .where(Conference.id != conference.id)
                .join(Participation, Participation.conference_id == Conference.id)
                .distinct()
                .order_by(Conference.edition_year.desc().nullslast())
                .limit(get_settings().brief_max_past_editions)
            )
        )
        .scalars()
        .all()
    )
    if not rows:
        return []

    people = (
        (
            await db.execute(
                select(Participation)
                .where(Participation.conference_id.in_([r.id for r in rows]))
                .order_by(Participation.person_label)
            )
        )
        .scalars()
        .all()
    )
    by_conference: dict[UUID, list[Participation]] = {}
    for row in people:
        by_conference.setdefault(row.conference_id, []).append(row)

    return [
        {
            "name": r.name,
            "year": r.edition_year,
            "verdict": r.attendance_verdict,
            "spend_usd": r.spend_usd,
            "audience_size_estimate": r.audience_size_estimate,
            "notes": r.attendance_notes,
            "attendees": [
                {
                    "sme_id": str(p.sme_id) if p.sme_id else None,
                    "full_name": p.person_label,
                    "activity": p.activity,
                    "outcome": p.outcome,
                }
                for p in by_conference.get(r.id, [])
            ],
        }
        for r in rows
    ]


async def _matched_pillar(db: AsyncSession, conference_id: UUID) -> dict | None:
    row = (
        await db.execute(
            select(StrategicPillar.name, StrategicPillar.description, ConferencePillar.score)
            .join(StrategicPillar, StrategicPillar.id == ConferencePillar.pillar_id)
            .where(ConferencePillar.conference_id == conference_id)
            .order_by(ConferencePillar.score.desc())
            .limit(1)
        )
    ).first()
    if row is None:
        return None
    return {"name": row[0], "description": row[1], "score": _round(row[2])}


async def _top_topics(db: AsyncSession, conference_id: UUID) -> list[dict]:
    """Topic strings straight off the conference row.

    Used to join the topic-vocabulary tables; those are gone, and the
    ``conferences.topics`` string array is the source of truth.
    """
    conf = await db.get(Conference, conference_id)
    names = list(conf.topics or []) if conf else []
    return [{"name": n} for n in names[: get_settings().brief_max_topics]]


async def _attendees(db: AsyncSession, conference_id: UUID) -> dict:
    """Who is actually going, or who actually went.

    This used to render a combinatorial optimiser's pick of the "best"
    team of 1, 2 or 3 SMEs, scored on coverage and redundancy. Nobody
    asked for that. The data model says track WHO IS GOING — a fact a
    person records, not an answer a machine computes — and until D11
    there was nowhere to record it, so an optimiser filled the gap.

    Now there is: participation rows carry the person, what they are
    doing (talk / booth / attend / sponsor), and when they travel.
    """
    rows = (
        (
            await db.execute(
                select(Participation)
                .where(Participation.conference_id == conference_id)
                .order_by(Participation.activity, Participation.person_label)
            )
        )
        .scalars()
        .all()
    )

    today = date.today()
    members = [
        {
            "person_label": r.person_label,
            "sme_id": str(r.sme_id) if r.sme_id else None,
            "activity": r.activity,
            "arrives_on": r.arrives_on.isoformat() if r.arrives_on else None,
            "departs_on": r.departs_on.isoformat() if r.departs_on else None,
            "has_attended": r.attended_at is not None
            or (r.departs_on is not None and r.departs_on < today),
            "notes": r.notes,
        }
        for r in rows
    ]
    return {
        "members": members,
        # Nobody recorded yet is a real and common state — a conference
        # approved but not yet staffed. Say so rather than showing an
        # invented recommendation.
        "source": "participation" if members else "none",
    }


def _cfp_section(conference: Conference) -> dict:
    deadlines = list(conference.cfp_deadlines or [])
    today = date.today()
    enriched: list[dict] = []
    next_idx: int | None = None
    next_days: int | None = None
    for i, d in enumerate(deadlines):
        date_str = d.get("date") if isinstance(d, dict) else None
        days_remaining: int | None = None
        if date_str:
            try:
                target = date.fromisoformat(date_str)
                days_remaining = (target - today).days
                if days_remaining >= 0 and (next_days is None or days_remaining < next_days):
                    next_days = days_remaining
                    next_idx = i
            except ValueError:
                pass
        enriched.append(
            {
                "kind": d.get("kind") if isinstance(d, dict) else None,
                "date": date_str,
                "description": d.get("description") if isinstance(d, dict) else None,
                "days_remaining": days_remaining,
                "is_next": False,
            }
        )
    if next_idx is not None:
        enriched[next_idx]["is_next"] = True
    return {
        "deadlines": enriched,
        "topics_of_interest": list(conference.cfp_topics_of_interest or [])[:get_settings().brief_max_topics],
        "open_at": _iso_date(conference.cfp_open_at),
        "close_at": _iso_date(conference.cfp_close_at),
    }


async def _talking_points(
    db: AsyncSession,
    conference: Conference,
) -> list[dict]:
    """Top messaging documents matched against this conference's text.

    One embedding call (the query is the conference name + topics + CFP
    topics of interest). Cached per conference by the outer
    ``build_brief``.
    """
    query_parts: list[str] = [conference.name]
    if conference.topics:
        query_parts.append(" ".join(conference.topics))
    if conference.cfp_topics_of_interest:
        query_parts.append(" ".join(conference.cfp_topics_of_interest))
    query = " ".join(p for p in query_parts if p).strip()
    if not query:
        return []

    try:
        chunks = await similar_chunks(
            db,
            query=query,
            owner_types=["messaging"],
            k=12,
            purpose="brief_talking_points",
            bump_last_used=False,
        )
    except Exception as exc:
        log.warning("brief.talking_points.embed_failed", error=str(exc))
        return []

    seen: dict[UUID, float] = {}
    for hit in chunks:
        prev = seen.get(hit.chunk.owner_id)
        if prev is None or hit.similarity > prev:
            seen[hit.chunk.owner_id] = hit.similarity
    if not seen:
        return []

    top_ids = sorted(seen, key=lambda i: seen[i], reverse=True)[:get_settings().brief_max_talking_docs]
    docs = (
        (await db.execute(select(MessagingDocument).where(MessagingDocument.id.in_(top_ids))))
        .scalars()
        .all()
    )
    by_id = {d.id: d for d in docs}

    out: list[dict] = []
    for did in top_ids:
        d = by_id.get(did)
        if d is None:
            continue
        out.append(
            {
                "document_id": str(d.id),
                "title": d.title,
                "elevator_pitch": d.elevator_pitch,
                "talking_points": (d.talking_points or [])[:get_settings().brief_max_talking_points_per_doc],
                "key_themes": (d.key_themes or [])[:get_settings().brief_max_talking_points_per_doc],
                "similarity": round(seen[did], 4),
            }
        )
    return out


async def _latest_decision(db: AsyncSession, conference_id: UUID) -> Decision | None:
    return (
        await db.execute(
            select(Decision)
            .where(Decision.conference_id == conference_id)
            .order_by(Decision.decided_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


async def _sources_count(db: AsyncSession, conference_id: UUID) -> int:
    return int(
        (
            await db.execute(
                select(func.count(ConferenceSource.raw_page_id)).where(
                    ConferenceSource.conference_id == conference_id
                )
            )
        ).scalar_one()
        or 0
    )


def _series_summary(series: ConferenceSeries, past: list[dict]) -> dict:
    attended_recent = sum(1 for p in past if p["attendees"])
    return {
        "id": str(series.id),
        "canonical_name": series.canonical_name,
        "typical_month": series.typical_month,
        "past_editions_count": len(past),
        "team_attended_recent": attended_recent,
    }


def _decision_summary(d: Decision | None) -> dict | None:
    if d is None:
        return None
    return {
        "decision": d.decision,
        "decided_by_label": d.decided_by_label,
        "decided_at": d.decided_at.isoformat() if d.decided_at else None,
        "reason": d.reason,
    }


def _bucket(score: float | None) -> str | None:
    if score is None:
        return None
    if score >= 0.75:
        return "strong"
    if score >= 0.55:
        return "good"
    if score >= 0.40:
        return "marginal"
    return "weak"


def _round(x: float | None) -> float | None:
    if x is None:
        return None
    return round(float(x), 4)


def _iso_date(d: date | None) -> str | None:
    return d.isoformat() if d else None


# ==========================================================================
# export.py — "Export view" on the conference list
# ==========================================================================
#
# The boss-facing spreadsheet. One row per conference in the CURRENT
# filtered view, every column the app knows about — including the ones
# nobody has filled in yet (actual spend, leads, worth-it verdict). Empty
# columns are the point: the export doubles as the checklist of what is
# still unrecorded.

#: (header, key) pairs — one place defines column order for BOTH formats.
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Rank", "rank"),
    ("Conference", "name"),
    ("Status", "status"),
    ("Event kind", "event_kind"),
    ("Overall score", "overall_score"),
    ("Strategy fit", "fit_score"),
    ("Speaker fit", "speaker_score"),
    ("Start date", "start_date"),
    ("End date", "end_date"),
    ("City", "location_city"),
    ("Country", "location_country"),
    ("Venue", "venue"),
    ("Virtual", "is_virtual"),
    ("Website", "website"),
    ("CFP URL", "cfp_url"),
    ("CFP opens", "cfp_open_at"),
    ("CFP closes", "cfp_close_at"),
    ("CFP topics of interest", "cfp_topics_of_interest"),
    ("Topics", "topics"),
    ("Acceptance rate %", "acceptance_rate_percent"),
    ("Audience size (est.)", "audience_size_estimate"),
    ("Estimated cost (USD)", "estimated_cost_usd"),
    ("Actual spend (USD)", "spend_usd"),
    ("Leads generated", "leads_generated"),
    ("Worth it?", "attendance_verdict"),
    ("Attendance notes", "attendance_notes"),
    ("Who is going", "who_is_going"),
    ("Previously attended", "previously_attended"),
    ("Travel logistics", "logistics_travel"),
    ("Lodging", "logistics_lodging"),
    ("Booth", "logistics_booth"),
    ("Sponsorship", "logistics_sponsorship"),
    ("Description", "description"),
    ("Added", "created_at"),
    ("Last updated", "updated_at"),
]

_EXPORT_SCORE_KEYS = {"overall_score", "fit_score", "speaker_score"}


def _export_cell(row: dict, key: str) -> Any:
    v = row.get(key)
    if v is None:
        return ""
    # Scores export as 0-100 like every screen shows them.
    if key in _EXPORT_SCORE_KEYS:
        return round(float(v) * 100)
    if isinstance(v, bool):
        return "yes" if v else "no"
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return v


def export_rows_to_csv(rows: list[dict]) -> bytes:
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for h, _ in EXPORT_COLUMNS])
    for row in rows:
        w.writerow([_export_cell(row, k) for _, k in EXPORT_COLUMNS])
    # BOM so Excel opens UTF-8 city names correctly on double-click.
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def export_rows_to_xlsx(rows: list[dict]) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Conferences"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EE0000")  # Red Hat red
    for col, (header, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(rows, start=2):
        for col, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(row=r, column=col, value=_export_cell(row, key))

    # Freeze the header, give it a filter, and size columns to content
    # (capped — Description would otherwise be a mile wide).
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, (header, key) in enumerate(EXPORT_COLUMNS, start=1):
        widest = len(header)
        for row in rows[:200]:
            widest = max(widest, len(str(_export_cell(row, key))))
        ws.column_dimensions[get_column_letter(col)].width = min(widest + 2, 50)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==========================================================================
# analytics.py — per-SME and per-pillar performance from participation
# ==========================================================================
#
# The "who is going" data (participation rows) plus the per-conference
# outcome fields (spend, leads, worth-it verdict) answer two questions the
# UI previously never asked: how is each PERSON doing on the circuit, and
# how is each PILLAR paying off. Both readers below are pure aggregation —
# no writes, no LLM — so the frontend renders numbers it never computes.
#
# Attribution honesty: spend/leads/worth-it live on the CONFERENCE row.
# For an SME they are the outcomes of events that person was part of, not
# a personal P&L; for a pillar, a conference aligned to two pillars counts
# toward both, so pillar sums can overlap and MUST NOT be totalled across
# pillars.


async def sme_analytics(db: AsyncSession, sme_id: UUID) -> dict:
    """Everything the SME performance panel shows, in one query pass."""
    rows = (
        await db.execute(
            select(Participation, Conference)
            .join(Conference, Conference.id == Participation.conference_id)
            .where(Participation.sme_id == sme_id)
            .order_by(Conference.start_date.desc().nulls_last())
        )
    ).all()

    today = date.today()
    by_activity: dict[str, int] = {}
    attended = 0
    upcoming = 0
    spend_total = 0
    leads_total = 0
    verdicts: dict[str, int] = {}
    events: list[dict] = []
    seen_conf_outcomes: set[UUID] = set()

    for p, c in rows:
        by_activity[p.activity] = by_activity.get(p.activity, 0) + 1
        went = p.attended_at is not None
        if went:
            attended += 1
        elif c.start_date is not None and c.start_date >= today:
            upcoming += 1
        # Outcome fields are event-level; count each conference once even
        # when this SME did two activities there.
        if went and c.id not in seen_conf_outcomes:
            seen_conf_outcomes.add(c.id)
            spend_total += c.spend_usd or 0
            leads_total += c.leads_generated or 0
            if c.attendance_verdict:
                verdicts[c.attendance_verdict] = verdicts.get(c.attendance_verdict, 0) + 1
        events.append(
            {
                "conference_id": str(c.id),
                "conference_name": c.name,
                "start_date": c.start_date.isoformat() if c.start_date else None,
                "activity": p.activity,
                "attended": went,
                "spend_usd": c.spend_usd,
                "leads_generated": c.leads_generated,
                "attendance_verdict": c.attendance_verdict,
            }
        )

    return {
        "sme_id": str(sme_id),
        "events_total": len(rows),
        "events_attended": attended,
        "events_upcoming": upcoming,
        "by_activity": by_activity,
        "attended_events_spend_usd": spend_total,
        "attended_events_leads": leads_total,
        "verdicts": verdicts,
        "events": events[:50],
    }


async def pillar_analytics(db: AsyncSession, pillar_id: UUID) -> dict:
    """Pillar performance: outcomes across conferences aligned to it.

    Alignment = a conference_pillars edge (matcher-written). A conference
    aligned to two pillars appears in both pillars' numbers — sums here
    are per-pillar views, not shares of one budget.
    """
    rows = (
        await db.execute(
            select(Conference)
            .join(ConferencePillar, ConferencePillar.conference_id == Conference.id)
            .where(ConferencePillar.pillar_id == pillar_id)
        )
    ).all()
    confs = [c for (c,) in rows]
    conf_ids = [c.id for c in confs]

    going_rows = (
        (
            await db.execute(
                select(Participation).where(Participation.conference_id.in_(conf_ids))
            )
        )
        .scalars()
        .all()
        if conf_ids
        else []
    )
    going_by_conf: dict[UUID, list[Participation]] = {}
    for p in going_rows:
        going_by_conf.setdefault(p.conference_id, []).append(p)

    today = date.today()
    attended_confs = 0
    planned_confs = 0
    participants = 0
    spend_total = 0
    leads_total = 0
    verdicts: dict[str, int] = {}
    outcomes: list[dict] = []

    for c in confs:
        people = going_by_conf.get(c.id, [])
        participants += len(people)
        went = any(p.attended_at is not None for p in people)
        planned = bool(people) and not went and (c.start_date is None or c.start_date >= today)
        if went:
            attended_confs += 1
            spend_total += c.spend_usd or 0
            leads_total += c.leads_generated or 0
            if c.attendance_verdict:
                verdicts[c.attendance_verdict] = verdicts.get(c.attendance_verdict, 0) + 1
            outcomes.append(
                {
                    "conference_id": str(c.id),
                    "conference_name": c.name,
                    "start_date": c.start_date.isoformat() if c.start_date else None,
                    "n_people": len(people),
                    "spend_usd": c.spend_usd,
                    "leads_generated": c.leads_generated,
                    "attendance_verdict": c.attendance_verdict,
                }
            )
        elif planned:
            planned_confs += 1

    return {
        "pillar_id": str(pillar_id),
        "conferences_aligned": len(confs),
        "conferences_attended": attended_confs,
        "conferences_planned": planned_confs,
        "participants_total": participants,
        "spend_usd_total": spend_total,
        "leads_total": leads_total,
        "cost_per_lead_usd": (
            round(spend_total / leads_total, 2) if spend_total and leads_total else None
        ),
        "verdicts": verdicts,
        "attended": outcomes[:50],
    }


# ==========================================================================
# analytics_charts.py — chart-ready series for the /analytics page
# ==========================================================================
#
# One endpoint, one pass, everything pre-binned server-side. The frontend
# draws axes and bars; it never aggregates. Filters narrow the conference
# set BEFORE binning so every chart answers the same filtered question.


async def analytics_overview(
    db: AsyncSession,
    *,
    pillar_id: UUID | None = None,
    country: str | None = None,
    months: int = 12,
    status: list[str] | None = None,
    event_kind: list[str] | None = None,
    include_virtual: bool = True,
    starts_after: date | None = None,
    starts_before: date | None = None,
) -> dict:
    """Every series the analytics page draws, filtered consistently."""
    stmt = select(Conference).where(Conference.status != "quarantined")
    if pillar_id is not None:
        stmt = stmt.join(
            ConferencePillar, ConferencePillar.conference_id == Conference.id
        ).where(ConferencePillar.pillar_id == pillar_id)
    if country:
        stmt = stmt.where(Conference.location_country == country.upper())
    if status:
        stmt = stmt.where(Conference.status.in_(status))
    if event_kind:
        stmt = stmt.where(Conference.event_kind.in_(event_kind))
    if not include_virtual:
        stmt = stmt.where(Conference.is_virtual.is_(False))
    if starts_after is not None:
        stmt = stmt.where(Conference.start_date >= starts_after)
    if starts_before is not None:
        stmt = stmt.where(Conference.start_date <= starts_before)
    confs = (await db.execute(stmt)).scalars().all()
    conf_ids = [c.id for c in confs]
    conf_by_id = {c.id: c for c in confs}

    # Latest match per conference for score histograms.
    match_rows = (
        (
            await db.execute(
                select(
                    Match.conference_id,
                    Match.overall_score,
                    Match.fit_score,
                    Match.speaker_score,
                    Match.judge_verdict,
                ).where(
                    Match.conference_id.in_(conf_ids)
                )
            )
        ).all()
        if conf_ids
        else []
    )

    # Participation for spend/leads/activity series.
    parts = (
        (
            await db.execute(
                select(Participation).where(Participation.conference_id.in_(conf_ids))
            )
        )
        .scalars()
        .all()
        if conf_ids
        else []
    )

    today = date.today()

    # ---- Status funnel ------------------------------------------------
    status_counts: dict[str, int] = {}
    for c in confs:
        status_counts[c.status] = status_counts.get(c.status, 0) + 1
    status_funnel = [
        {"status": s, "count": n}
        for s, n in sorted(status_counts.items(), key=lambda kv: -kv[1])
    ]

    # ---- Overall-score histogram (bins of 10, 0-100) ------------------
    bins = [0] * 10
    for _cid, overall, _fit, _spk, _jv in match_rows:
        if overall is None:
            continue
        b = min(int(float(overall) * 100) // 10, 9)
        bins[b] += 1
    score_histogram = [
        {"bucket": f"{i * 10}-{i * 10 + 9}", "count": n} for i, n in enumerate(bins)
    ]

    # ---- CFP deadlines by month (next `months`) -----------------------
    def _month_key(d: date) -> str:
        return f"{d.year}-{d.month:02d}"

    month_keys: list[str] = []
    y, m = today.year, today.month
    for _ in range(months):
        month_keys.append(f"{y}-{m:02d}")
        m += 1
        if m == 13:
            y, m = y + 1, 1
    cfp_by_month = dict.fromkeys(month_keys, 0)
    for c in confs:
        if c.cfp_close_at and c.cfp_close_at >= today:
            k = _month_key(c.cfp_close_at)
            if k in cfp_by_month:
                cfp_by_month[k] += 1
    cfp_months = [{"month": k, "count": v} for k, v in cfp_by_month.items()]

    # ---- Conferences by country (top 12) ------------------------------
    country_counts: dict[str, int] = {}
    for c in confs:
        key = "virtual" if c.is_virtual else (c.location_country or "unknown")
        country_counts[key] = country_counts.get(key, 0) + 1
    by_country = [
        {"country": k, "count": n}
        for k, n in sorted(country_counts.items(), key=lambda kv: -kv[1])[:12]
    ]

    # ---- Outcomes by month (spend / leads over trailing `months`) -----
    trailing: list[str] = []
    y, m = today.year, today.month
    for _ in range(months):
        trailing.append(f"{y}-{m:02d}")
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    trailing.reverse()
    spend_by_month = dict.fromkeys(trailing, 0)
    leads_by_month = dict.fromkeys(trailing, 0)
    attended_conf_ids = {p.conference_id for p in parts if p.attended_at is not None}
    for cid in attended_conf_ids:
        c = conf_by_id.get(cid)
        if c is None or c.start_date is None:
            continue
        k = _month_key(c.start_date)
        if k in spend_by_month:
            spend_by_month[k] += c.spend_usd or 0
            leads_by_month[k] += c.leads_generated or 0
    outcomes_by_month = [
        {"month": k, "spend_usd": spend_by_month[k], "leads": leads_by_month[k]}
        for k in trailing
    ]

    # ---- Activity mix -------------------------------------------------
    activity_counts: dict[str, int] = {}
    for p in parts:
        activity_counts[p.activity] = activity_counts.get(p.activity, 0) + 1
    activity_mix = [{"activity": a, "count": n} for a, n in activity_counts.items()]

    # ---- Matcher signals ---------------------------------------------
    # What the algorithm actually saw for this view: signal averages,
    # judge verdicts, and the live weights/gates so the reader can tell
    # WHY the funnel looks the way it does.
    fits = [float(f) for _c, _o, f, _s, _j in match_rows if f is not None]
    spks = [float(s) for _c, _o, _f, s, _j in match_rows if s is not None]
    judge_counts: dict[str, int] = {}
    for _c, _o, _f, _s, jv in match_rows:
        key = jv or "no_verdict"
        judge_counts[key] = judge_counts.get(key, 0) + 1
    s = get_settings()
    matcher_signals = {
        "scored": len(match_rows),
        "avg_fit": round(sum(fits) / len(fits), 4) if fits else None,
        "avg_speakers": round(sum(spks) / len(spks), 4) if spks else None,
        "judge_verdicts": [
            {"verdict": k, "count": n} for k, n in sorted(judge_counts.items())
        ],
        "weights": {"fit": s.match_w_fit, "speakers": s.match_w_speakers},
        "gates": {"messaging": s.match_m_gate, "speakers": s.match_s_gate},
        "sme_dimension_weights": {
            "audience": s.sme_w_audience,
            "bio": s.sme_w_bio,
            "location": s.sme_w_location,
            "past": s.sme_w_past,
        },
    }

    # ---- Talks library ------------------------------------------------
    talk_rows = (
        await db.execute(
            select(Talk.review_status, func.count(Talk.id))
            .where(Talk.is_active.is_(True))
            .group_by(Talk.review_status)
        )
    ).all()
    sub_rows = (
        await db.execute(
            select(TalkSubmission.outcome, func.count(TalkSubmission.id)).group_by(
                TalkSubmission.outcome
            )
        )
    ).all()
    talks = {
        "active_total": sum(int(n) for _st, n in talk_rows),
        "by_review_status": [
            {"status": st, "count": int(n)} for st, n in talk_rows
        ],
        "submissions_by_outcome": [
            {"outcome": o or "pending", "count": int(n)} for o, n in sub_rows
        ],
    }

    # ---- SME roster ---------------------------------------------------
    sme_rows = (
        await db.execute(
            select(Sme.id, Sme.full_name, Sme.expertise).where(Sme.is_active.is_(True))
        )
    ).all()
    # Everyone who has been to anything — roster SMEs by id, guests and
    # not-yet-registered colleagues by their free-text label. The team's
    # reality includes people who are not SME records, and hiding them
    # made the imported history look like nobody went anywhere.
    person_rows = (
        await db.execute(
            select(
                Participation.sme_id,
                Participation.person_label,
                func.count(Participation.id),
            ).group_by(Participation.sme_id, Participation.person_label)
        )
    ).all()
    sme_name_by_id = {sid: name for sid, name, _e in sme_rows}
    events_by_person: dict[str, dict] = {}
    for sid, label, n in person_rows:
        if sid is not None and sid in sme_name_by_id:
            key, on_roster = sme_name_by_id[sid], True
        else:
            key, on_roster = (label or "Unknown").strip(), False
        entry = events_by_person.setdefault(
            key, {"name": key, "events": 0, "on_roster": on_roster}
        )
        entry["events"] += int(n)
        entry["on_roster"] = entry["on_roster"] or on_roster
    # Roster members with zero events still show — coverage gaps matter.
    for sid, name, _e in sme_rows:
        events_by_person.setdefault(
            name, {"name": name, "events": 0, "on_roster": True}
        )
    smes = {
        "active_total": len(sme_rows),
        "with_expertise": sum(1 for _i, _n, e in sme_rows if (e or "").strip()),
        "events_per_sme": sorted(
            events_by_person.values(), key=lambda r: -r["events"]
        )[:20],
    }

    # ---- Pillar alignment --------------------------------------------
    pillar_rows = (
        await db.execute(
            select(
                StrategicPillar.name,
                func.count(ConferencePillar.conference_id),
                func.avg(ConferencePillar.score),
            )
            .join(ConferencePillar, ConferencePillar.pillar_id == StrategicPillar.id)
            .group_by(StrategicPillar.name)
        )
    ).all()
    pillar_alignment = [
        {"pillar": name, "conferences": int(n), "avg_score": round(float(avg), 4)}
        for name, n, avg in pillar_rows
    ]

    return {
        "filters": {
            "pillar_id": str(pillar_id) if pillar_id else None,
            "country": country.upper() if country else None,
            "months": months,
        },
        "matcher_signals": matcher_signals,
        "talks": talks,
        "smes": smes,
        "pillar_alignment": pillar_alignment,
        "conference_count": len(confs),
        "status_funnel": status_funnel,
        "score_histogram": score_histogram,
        "cfp_deadlines_by_month": cfp_months,
        "by_country": by_country,
        "outcomes_by_month": outcomes_by_month,
        "activity_mix": activity_mix,
        "totals": {
            "attended": len(attended_conf_ids),
            "spend_usd": sum(v["spend_usd"] for v in outcomes_by_month),
            "leads": sum(v["leads"] for v in outcomes_by_month),
        },
    }
