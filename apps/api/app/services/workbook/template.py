"""Empty workbook template generator (plan 31).

Builds a Google-Sheets-friendly XLSX:

  * A **README** sheet (first tab) that explains what the workbook is for,
    what you can and can't change, and how to upload it back.
  * One sheet per ``SheetSpec`` with:
      - Row 1: column headers, with cell **comments** for per-column
        guidance (visible on hover in both Excel and Google Sheets).
      - Row 2: a sample row marked "Sample —" that the reader skips on
        import (so the user doesn't have to remember to delete it).
      - Enum columns get **data validation** so the cell becomes a
        dropdown picker.
      - Required columns get a different header fill from optional ones,
        and system columns (``_scout_id`` / ``_action``) are visually
        distinct from real data columns.

The reader (``apps/api/app/services/workbook/reader.py``) is the source of
truth for what's parseable — anything cosmetic added here must leave row 1
as the header and row 2+ as the data range.

Used by ``GET /api/v1/config/workbook-template``.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.services.workbook._schema import SHEET_SPECS, ColumnSpec, SheetSpec

# Palette — picked so headers are legible in both Excel and Google Sheets.
_REQUIRED_FILL = PatternFill("solid", fgColor="B91C1C")  # red-700 — "you must fill this"
_OPTIONAL_FILL = PatternFill("solid", fgColor="374151")  # gray-700 — "fill if you have it"
_SYSTEM_FILL = PatternFill("solid", fgColor="64748B")  # slate-500 — managed by Scout
_HEADER_FONT = Font(bold=True, color="F9FAFB", size=11)
_SAMPLE_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100 — "I'm a sample, delete me"
_SAMPLE_FONT = Font(italic=True, color="92400E")  # amber-800
_THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Names of the synthetic system columns — styled differently.
_SYSTEM_COLUMNS = {"_scout_id", "_action"}


def build_empty_template() -> bytes:
    wb = Workbook()
    # Default sheet from Workbook() — repurpose as README.
    readme = wb.active
    readme.title = "README"
    _build_readme_sheet(readme)

    for spec in SHEET_SPECS:
        ws = wb.create_sheet(spec.name)
        _write_header(ws, spec)
        if spec.name == "Settings":
            _write_settings_rows(ws, spec)
        else:
            _write_sample_row(ws, spec)
        _attach_validation(ws, spec)
        _set_column_widths(ws, spec)
        ws.sheet_view.showGridLines = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_settings_rows(ws: Worksheet, spec: SheetSpec) -> None:
    """Pre-populate the Settings tab with all 41 settings + their default
    metadata. `value` cells are left BLANK for secrets (operator fills in
    LLM API key) and for everything else the cell shows the env-default
    so the operator can see what's currently in effect. Edit any cell to
    set an override; leave blank to inherit."""
    from app.services.workbook.writer import _settings_rows

    rows = _settings_rows(populate_with_current=False)
    for r in rows:
        ws.append([r.get(col.name, "") for col in spec.columns])

    # Style metadata columns (name, kind, group, description,
    # restart_required, is_secret) so operators see they're read-only.
    # Only the `value` column is operator-editable.
    from openpyxl.styles import Font, PatternFill

    metadata_fill = PatternFill("solid", fgColor="F3F4F6")  # gray-100
    metadata_font = Font(color="6B7280", size=10)  # gray-500
    secret_value_fill = PatternFill("solid", fgColor="FEE2E2")  # red-100

    for row_idx in range(2, ws.max_row + 1):
        # Find the column indexes once.
        col_by_name = {col.name: i + 1 for i, col in enumerate(spec.columns)}
        is_secret_idx = col_by_name.get("is_secret")
        if is_secret_idx and ws.cell(row=row_idx, column=is_secret_idx).value == "TRUE":
            # Highlight the value cell red for secret rows so the API-key
            # field jumps out visually.
            value_idx = col_by_name["value"]
            ws.cell(row=row_idx, column=value_idx).fill = secret_value_fill

        # De-emphasize the metadata cells.
        for col_name in ("name", "kind", "group", "description", "restart_required", "is_secret"):
            idx = col_by_name.get(col_name)
            if idx:
                cell = ws.cell(row=row_idx, column=idx)
                cell.fill = metadata_fill
                cell.font = metadata_font


# ---------------------------------------------------------------------------
# README sheet — the friendly entry point
# ---------------------------------------------------------------------------
def _build_readme_sheet(ws: Worksheet) -> None:
    """The README is the operator's primary onboarding doc. It's the first
    tab so it's what you see when you open the file. Goal: a fresh user
    can fill in this workbook + run a fresh install without ever reading
    the GitHub repo's docs.
    """
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100
    ws.sheet_view.showGridLines = False

    # --- Hero ----------------------------------------------------------
    ws.merge_cells("A1:B1")
    ws["A1"] = "Scout — Configuration Workbook"
    ws["A1"].font = Font(bold=True, size=22, color="111827")

    ws.merge_cells("A2:B2")
    ws["A2"] = (
        "One file. Everything Scout needs to run: your LLM API keys, "
        "your matcher tuning, and the team data (SMEs, audiences, "
        "messaging, pillars). Fill it in once, upload at Settings → "
        "Workbook, and you have a working install."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = Font(size=12, color="374151")
    ws.row_dimensions[2].height = 60

    # Track our current row so adding/removing sections is a one-line edit.
    row = 4

    # --- Quick install -------------------------------------------------
    row = _readme_section(ws, row, "Quick install — clone to running in ~5 minutes")
    row = _readme_numbered(
        ws,
        row,
        [
            (
                "Install Docker Desktop or Podman + podman-compose. "
                "Anything that runs containers will work."
            ),
            (
                "Clone the repo: git clone "
                "https://github.com/<your-org>/scout (or your fork)."
            ),
            (
                "Copy .env.example to .env. Open .env in a text editor and "
                "paste your LLM API key into the LLM_API_KEY field. "
                "If you don't have a key yet, leave LLM_DRY_RUN=true and "
                "Scout will boot with canned LLM responses so you can poke "
                "the UI offline."
            ),
            ("Run `make up` (~2 min first time). Then `make migrate` (~5 sec)."),
            (
                "Open http://localhost:8000 in your browser. You should see "
                "the Scout dashboard with empty stats — that's correct, you "
                "haven't loaded any data yet."
            ),
            (
                "Go to Settings → Workbook → Import. Upload THIS file. Scout "
                "will preview what will change. Click Apply. Done."
            ),
        ],
    )
    row += 1

    # --- Step 1 of fill-in: LLM API key (most important) --------------
    row = _readme_section(ws, row, "Step 1 — Your LLM API keys (Settings tab)")
    row = _readme_para(
        ws,
        row,
        (
            "Open the Settings tab. Look for the row where name = "
            "'llm_api_key'. The value cell is highlighted in red — that's "
            "the cell you NEED to fill in for Scout to do anything useful. "
            "Paste your chat-model API key there. If your embedding "
            "model uses a separate key (different scope), also fill in "
            "llm_embedding_api_key.\n\n"
            "Where to get a key: log into your LLM provider's dashboard "
            "and create a token with access to a chat model and an "
            "embedding model (Scout defaults expect nomic-embed-text-v1-5 "
            "for embeddings). Set llm_base_url to your provider's "
            "OpenAI-compatible endpoint."
        ),
    )
    row += 1

    # --- Step 2: team data ---------------------------------------------
    row = _readme_section(ws, row, "Step 2 — Your team data")
    row = _readme_para(
        ws,
        row,
        (
            "These tabs tell Scout who you are and what you care about. The "
            "minimum-to-be-useful is:"
        ),
    )
    row = _readme_bullets(
        ws,
        row,
        [
            (
                "At least one Audience (so the matcher can compute "
                "audience overlap). Try one persona you actually pitch to."
            ),
            (
                "At least one SME with a real, 200+ character bio and 2+ "
                "topic assignments. The matcher embeds the bio for "
                "similarity, so 'Sarah works on AI' is useless — write the "
                "kind of paragraph you'd put on her speaker page."
            ),
            (
                "At least one Topic (or 5-10 — the matcher uses topic "
                "overlap heavily)."
            ),
            (
                "Pillars and Series are useful but optional for first run. "
                "Series only matters once you have past conferences logged."
            ),
        ],
    )
    row = _readme_para(
        ws,
        row,
        (
            "Once you have one row in each of those four tabs, Scout can "
            "score events. More data = better recommendations."
        ),
    )
    row += 1

    # --- Step 3: upload -----------------------------------------------
    row = _readme_section(ws, row, "Step 3 — Upload and apply")
    row = _readme_numbered(
        ws,
        row,
        [
            "Save the workbook as .xlsx (File → Download → Excel in Google Sheets).",
            "In Scout: Settings → Workbook → drop the file into the upload area.",
            (
                "Scout runs a PREVIEW first. It shows what would change in each "
                "tab: how many new rows, how many updates, how many deletes "
                "(if any). Nothing is written yet."
            ),
            (
                "Review the preview. If it looks right, click Apply. The whole "
                "import is one database transaction — either everything lands "
                "or nothing changes."
            ),
        ],
    )
    row += 1

    # --- Tabs at a glance ----------------------------------------------
    row = _readme_section(ws, row, "The 7 tabs in this workbook")
    tab_descriptions = [
        ("Settings", (
            "Every runtime tunable in one place: LLM keys, matcher "
            "weights, gate thresholds, AI keyword filter, log level. Edit "
            "the value column; leave blank to keep the current value. "
            "Rows with red value cells are SECRETS — handle the file "
            "carefully."
        )),
        ("Pillars", (
            "Your AI strategy's pillars. Edit display_order to "
            "control how they're listed on the dashboard. You probably "
            "have 3-5 of these."
        )),
        ("Industries", (
            "Allowed values for the Audiences.industry column. Adding an "
            "industry here lets a new audience row use it. Single column."
        )),
        ("Audiences", (
            "Who you're trying to reach at conferences (personas: "
            "Platform Engineering Lead, ML Ops Director, C-Suite, etc.). "
            "Each persona has pain points and key messages that the "
            "matcher uses to predict event fit."
        )),
        ("SMEs", (
            "Your subject-matter experts. Bio quality matters — the "
            "matcher embeds the bio text for similarity scoring. "
            "primary_topics and audience_focus reference Topics and "
            "Audiences by NAME (not UUID) and get resolved on import."
        )),
        ("Topics", (
            "Your controlled topic vocabulary. Adding a topic here "
            "promotes it to 'active' so the matcher uses it. Aliases let "
            "you collapse 'NeurIPS' and 'Neural Information Processing "
            "Systems' to the same topic."
        )),
        ("Series", (
            "Year-over-year conference series. If you've been to PyCon "
            "2024, this tab is what lets Scout know PyCon 2026 is the "
            "next edition (and apply the past-attendance bonus to "
            "matching)."
        )),
    ]
    for name, desc in tab_descriptions:
        ws.cell(row=row, column=1, value=name).font = Font(
            bold=True, color="2563EB", size=11
        )
        ws.cell(row=row, column=2, value=desc).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = max(38, (len(desc) // 90 + 1) * 18)
        row += 1
    row += 1

    # --- The Settings tab in detail ------------------------------------
    row = _readme_section(ws, row, "The Settings tab — groups explained")
    row = _readme_para(
        ws,
        row,
        (
            "The Settings tab has 41 rows. Don't be intimidated — most have "
            "sensible defaults. Here's what each group of settings controls:"
        ),
    )
    setting_groups = [
        ("llm", (
            "LLM connection: base URL, API key, chat + embedding model "
            "names, dry-run toggle, monthly USD budget cap. THIS IS WHERE "
            "YOUR API KEY GOES."
        )),
        ("matcher", (
            "How the fit score is computed. m_gate / p_gate / s_gate set "
            "the minimum score for each stage to count. w_messaging / "
            "w_pillar / w_sme are the weights — they should sum to 1.0."
        )),
        ("sme", (
            "How the per-SME composite score is built: weights for topic "
            "overlap, audience overlap, bio similarity, location, and "
            "past attendance."
        )),
        ("team", (
            "Multi-person team recommendation: how many candidates to "
            "consider, weights for individual fit vs coverage vs "
            "redundancy avoidance."
        )),
        ("decay", (
            "Whether old data 'fades' over time (Ebbinghaus-style)."
        )),
        ("discovery", (
            "Where Scout looks for events. Includes the multilingual AI "
            "keyword filter (148 terms by default), the search provider "
            "(DDG/Brave/Tavily) and any optional API keys, seed URLs to "
            "always crawl, URL blocklist, and the cron schedule."
        )),
        ("scraper", (
            "Default politeness delay and User-Agent for outbound HTTP."
        )),
        ("logging", (
            "Log level + format (json for production log shippers, "
            "console for human reading)."
        )),
    ]
    for group, desc in setting_groups:
        ws.cell(row=row, column=1, value=group).font = Font(
            bold=True, color="0F766E", size=11
        )
        ws.cell(row=row, column=2, value=desc).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = max(38, (len(desc) // 90 + 1) * 18)
        row += 1
    row += 1

    # --- Cell types ----------------------------------------------------
    row = _readme_section(ws, row, "Cell types you'll see")
    type_rows = [
        ("Text", "Plain text, e.g. \"Platform Engineering Leaders\"."),
        ("List of text", (
            "Multiple values in one cell, separated by semicolons. e.g. "
            "\"RAG; Embeddings; Vector DBs\". No trailing semicolon."
        )),
        ("Boolean", "TRUE or FALSE (case doesn't matter). Settings tab uses dropdowns."),
        ("Number", "Whole (1, 42) or decimal (0.5, 1.0)."),
        ("Date", "YYYY-MM-DD, e.g. 2026-09-15."),
        ("Dropdown", (
            "Click the cell — a dropdown arrow appears with the valid "
            "values. Used for enums like role_seniority, _action, log_level."
        )),
        ("ISO country", "Two letters, uppercase, e.g. US, DE, JP."),
        ("Secret (API key)", (
            "Plain text, but treated as sensitive — file should be "
            "chmod 600 and never committed to git."
        )),
    ]
    for label, body in type_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=row, column=2, value=body).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = max(26, (len(body) // 95 + 1) * 18)
        row += 1
    row += 1

    # --- Colour key ----------------------------------------------------
    row = _readme_section(ws, row, "Header & cell colour key")
    key_rows = [
        ("Red header", "Required column — the importer rejects the row if blank.", _REQUIRED_FILL),
        ("Dark grey header", "Optional column — fill in if you have the data.", _OPTIONAL_FILL),
        ("Slate header", "System column (_scout_id, _action) — leave alone.", _SYSTEM_FILL),
        ("Light grey cell", "Metadata column on Settings tab — read-only, do not edit.", PatternFill("solid", fgColor="F3F4F6")),
        ("Red value cell", "Secret (API key) on Settings tab — fill in carefully.", PatternFill("solid", fgColor="FEE2E2")),
        ("Amber row", "Sample row — replace or leave; the importer skips it.", _SAMPLE_FILL),
    ]
    for label, body, fill in key_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.fill = fill
        if fill in (_REQUIRED_FILL, _OPTIONAL_FILL, _SYSTEM_FILL):
            c1.font = Font(bold=True, color="F9FAFB")
        ws.cell(row=row, column=2, value=body).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1
    row += 1

    # --- What gets backed up here ---------------------------------------
    row = _readme_section(ws, row, "What this workbook DOES back up")
    row = _readme_bullets(
        ws,
        row,
        [
            "All settings: LLM API keys, matcher weights, AI keyword filter, etc.",
            "All your reference data: pillars, audiences, SMEs, topics, series.",
            (
                "Enough to bring a fresh install back to your operational "
                "state in one upload."
            ),
        ],
    )

    row = _readme_section(ws, row, "What this workbook does NOT back up")
    row = _readme_bullets(
        ws,
        row,
        [
            (
                "Conferences themselves. These are re-discovered from the "
                "live web feed each install — a feature, not a bug. You "
                "get this week's events, not last month's."
            ),
            (
                "Match scores, brief PDFs, decisions, agent chat history. "
                "These are derived from the data + settings above and get "
                "recomputed automatically."
            ),
            (
                "Past conferences (the log of what your team has attended). "
                "Not in the workbook yet; manually re-enter via the "
                "/past-conferences page after restore."
            ),
            "Messaging documents. Also manual re-entry via /messaging after restore.",
        ],
    )
    row += 1

    # --- Restore / move install ----------------------------------------
    row = _readme_section(ws, row, "Restore from this workbook (fresh install or new machine)")
    row = _readme_numbered(
        ws,
        row,
        [
            "Save your most recent workbook export somewhere safe.",
            (
                "If moving machines: copy this workbook + your .env file to the "
                "new machine."
            ),
            (
                "On the new machine, run the quick-install steps above (clone, "
                ".env, make up, make migrate)."
            ),
            (
                "Open http://localhost:8000/settings, upload the workbook, "
                "Preview, Apply."
            ),
            (
                "Click Discover more on /conferences to re-pull this week's "
                "events from the live feed."
            ),
            (
                "On first open of any conference detail or brief page, the "
                "matcher auto-runs (5-30 sec, shows a skeleton). You're back "
                "in business."
            ),
        ],
    )
    row += 1

    # --- Common mistakes -----------------------------------------------
    row = _readme_section(ws, row, "Common mistakes (and how to avoid them)")
    row = _readme_bullets(
        ws,
        row,
        [
            (
                "Renaming a column header. The importer looks columns up by "
                "name; rename = importer can't find it = row error. Don't "
                "edit row 1 of any tab."
            ),
            (
                "Renaming a tab. Same issue. The 7 tab names are wired into "
                "the importer."
            ),
            (
                "Adding a formula. Cells starting with =, +, -, @ are "
                "rejected on import (formula injection defense)."
            ),
            (
                "Wrong country code. ISO-3166-1 alpha-2 — TWO letters, "
                "uppercase. 'United States' = wrong, 'US' = right."
            ),
            (
                "Forgetting the semicolon separator. Lists are 'A; B; C' "
                "with semicolons + spaces, not 'A, B, C' commas."
            ),
            (
                "Trying to delete a setting by clearing the value. Blank "
                "value = 'leave alone'. To reset a setting to its env "
                "default, use the /settings/tunables UI's reset button."
            ),
            (
                "Importing the workbook before .env has a valid LLM API key. "
                "Scout's startup validator rejects the placeholder. Either "
                "set LLM_DRY_RUN=true in .env, or put your real key in .env "
                "before running 'make up'."
            ),
        ],
    )
    row += 1

    # --- Safety / final notes -------------------------------------------
    row = _readme_section(ws, row, "Safety notes")
    row = _readme_bullets(
        ws,
        row,
        [
            (
                "Scout always previews changes before writing. You can cancel."
            ),
            (
                "Rows in the database that are MISSING from your upload are "
                "KEPT, not deleted. The only way to delete is to explicitly "
                "set _action=delete on the row (soft-delete; row stays in DB "
                "with is_active=false)."
            ),
            (
                "This file contains your LLM API keys IN PLAIN TEXT. Save "
                "with chmod 600, store on encrypted disk, never commit to "
                "git, never share in Slack. Treat it like a password file."
            ),
            (
                "Need help? See docs/ops/runbook.md in the repo, or the "
                "Documentation section of the project README."
            ),
        ],
    )


def _readme_section(ws: Worksheet, row: int, title: str) -> int:
    """Write a section heading and return the next free row."""
    ws.cell(row=row, column=1, value=title).font = Font(
        bold=True, size=14, color="111827"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    return row + 1


def _readme_para(ws: Worksheet, row: int, body: str) -> int:
    """Write a paragraph (handles embedded \\n line breaks) and return
    the next free row. Height auto-scales to fit the wrapped text."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=body)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Approximation: assume ~110 chars per visual line at our column width,
    # plus account for explicit \n splits.
    visual_lines = 0
    for chunk in body.split("\n"):
        visual_lines += max(1, len(chunk) // 110 + (1 if len(chunk) % 110 else 0))
    ws.row_dimensions[row].height = max(20, visual_lines * 18)
    return row + 1


def _readme_bullets(ws: Worksheet, start_row: int, items: list[str]) -> int:
    """Write a bullet list and return the next free row."""
    for i, item in enumerate(items):
        row = start_row + i
        ws.cell(row=row, column=1, value="•").alignment = Alignment(
            horizontal="right", vertical="top"
        )
        ws.cell(row=row, column=2, value=item).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = max(20, (len(item) // 95 + 1) * 18)
    return start_row + len(items)


def _readme_numbered(ws: Worksheet, start_row: int, items: list[str]) -> int:
    """Write a numbered list — same as bullets but with '1.' / '2.' /…
    in the first column. Returns the next free row."""
    for i, item in enumerate(items):
        row = start_row + i
        ws.cell(row=row, column=1, value=f"{i + 1}.").alignment = Alignment(
            horizontal="right", vertical="top"
        )
        ws.cell(row=row, column=1).font = Font(bold=True, color="111827")
        ws.cell(row=row, column=2, value=item).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = max(20, (len(item) // 95 + 1) * 18)
    return start_row + len(items)


# ---------------------------------------------------------------------------
# Per-sheet header + sample row
# ---------------------------------------------------------------------------
def _write_header(ws: Worksheet, spec: SheetSpec) -> None:
    headers = spec.column_names()
    ws.append(headers)

    for col_idx, col_spec in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        if col_spec.name in _SYSTEM_COLUMNS:
            cell.fill = _SYSTEM_FILL
        elif col_spec.required:
            cell.fill = _REQUIRED_FILL
        else:
            cell.fill = _OPTIONAL_FILL

        cell.comment = _header_comment_for(col_spec)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _header_comment_for(col_spec: ColumnSpec) -> Comment:
    """Build a per-column hover comment summarizing required-ness, type,
    and any spec-level note (enum values, length cap, etc.)."""
    lines: list[str] = []
    if col_spec.name in _SYSTEM_COLUMNS:
        lines.append(f"SYSTEM COLUMN — {col_spec.name}")
        lines.append("Managed by Scout. Leave blank on new rows.")
    elif col_spec.required:
        lines.append(f"REQUIRED — {col_spec.name}")
    else:
        lines.append(f"Optional — {col_spec.name}")

    lines.append(f"Type: {_human_kind(col_spec.kind)}")
    if col_spec.enum_values:
        lines.append("Allowed values:")
        for v in col_spec.enum_values:
            lines.append(f"  • {v}")
    if col_spec.max_len is not None:
        lines.append(f"Max length: {col_spec.max_len} characters")
    if col_spec.note:
        lines.append("")
        lines.append(col_spec.note)

    # openpyxl Comment.width / .height are in points. ~260x140 fits ~7 lines.
    comment = Comment("\n".join(lines), author="Scout")
    comment.width = 280
    comment.height = max(120, 18 * (len(lines) + 1))
    return comment


def _human_kind(kind: str) -> str:
    """Convert the internal ColumnKind to something a human reading the
    cell-comment will recognize."""
    return {
        "text": "Text",
        "long_text": "Text (long-form, multi-line OK)",
        "int": "Whole number",
        "float": "Decimal number",
        "bool": "TRUE or FALSE",
        "date": "Date (YYYY-MM-DD)",
        "uuid": "UUID — leave blank on new rows",
        "list_text": "List of text, semicolon-separated (e.g. 'A; B; C')",
        "list_uuid": "List of UUIDs, semicolon-separated",
        "enum": "Pick one from the dropdown",
        "action": "upsert / delete / skip",
    }.get(kind, kind)


def _write_sample_row(ws: Worksheet, spec: SheetSpec) -> None:
    sample = _SAMPLE_ROWS.get(spec.name)
    if not sample:
        return
    ordered = [sample.get(c.name, "") for c in spec.columns]
    ws.append(ordered)

    # Style the sample row distinctively + comment the first cell.
    sample_row = ws.max_row
    for col_idx in range(1, len(spec.columns) + 1):
        cell = ws.cell(row=sample_row, column=col_idx)
        cell.fill = _SAMPLE_FILL
        cell.font = _SAMPLE_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    first_data_cell = ws.cell(row=sample_row, column=1)
    first_data_cell.comment = Comment(
        "This is a sample row. Replace it with real data, or leave it — "
        "the importer skips any row whose primary name starts with "
        "'Sample —' so you don't have to remember to delete it.",
        author="Scout",
    )


def _attach_validation(ws: Worksheet, spec: SheetSpec) -> None:
    """Add Excel/Google-Sheets data validation for enum and bool columns."""
    for col_idx, col_spec in enumerate(spec.columns, start=1):
        if col_spec.enum_values:
            values = list(col_spec.enum_values)
        elif col_spec.kind == "bool":
            values = ["TRUE", "FALSE"]
        else:
            continue

        # Quote values so commas inside any value don't break the formula.
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Pick one of: {', '.join(values)}",
        )
        col_letter = get_column_letter(col_idx)
        # Cover rows 2-1000 so the dropdown works as the user adds rows.
        dv.add(f"{col_letter}2:{col_letter}1000")
        ws.add_data_validation(dv)


_SAMPLE_ROWS: dict[str, dict[str, object]] = {
    "Pillars": {
        "name": "Sample — Trusted Open Source AI",
        "description": "Pillar focused on enterprise-grade open AI tooling.",
        "display_order": 1,
    },
    "Industries": {
        "name": "Sample — Tech",
    },
    "Audiences": {
        "name": "Sample — Platform Engineering Leaders",
        "industry": "Tech",
        "role_seniority": "director",
        "description": (
            "Sample audience: leaders responsible for the developer "
            "infrastructure and AI runtime platforms at midsize-to-large "
            "enterprises."
        ),
        "primary_pain_points": "Platform sprawl; Hard to measure platform ROI",
        "key_messages": "Open source by default; Standardize on Kubernetes",
        "exclusion_criteria": "",
        "is_active": "TRUE",
    },
    "SMEs": {
        "full_name": "Sample — Alice Chen",
        "email": "alice@example.com",
        "team": "Marketing",
        "expertise_areas": "Retrieval-augmented generation; Vector databases",
        "primary_topics": "rag; llm",
        "audience_focus": "Platform Engineering Leaders",
        "location_country": "US",
        "location_city": "Boston",
        "bio": (
            "Sample bio: 10+ years platform engineering, last three on "
            "retrieval-augmented generation systems at scale. Frequent "
            "speaker on RAG architecture, vector databases, and inference "
            "operationalization for enterprise platforms."
        ),
        "linkedin_url": "https://www.linkedin.com/in/example",
        "github_url": "",
        "website_url": "",
        "is_active": "TRUE",
    },
    "Topics": {
        "name": "Sample — RAG",
        "slug": "rag",
        "aliases": "Retrieval-Augmented Generation",
        "is_active": "TRUE",
        "pending_review": "FALSE",
    },
    "Series": {
        "canonical_name": "Sample — Made-Up Conference",
        "aliases": "MUC; My Mock Conf",
        "description": "Example series row; replace or delete.",
        "typical_month": 9,
        "typical_topics": "fictitious; sample",
        "homepage": "https://example.invalid",
        "is_active": "TRUE",
    },
}


def _set_column_widths(ws: Worksheet, spec: SheetSpec) -> None:
    """Starting widths so the template opens cleanly in Sheets."""
    widths = {
        "_scout_id": 36,
        "_action": 10,
        "name": 30,
        "full_name": 28,
        "canonical_name": 28,
        "description": 60,
        "bio": 70,
        "primary_pain_points": 50,
        "key_messages": 50,
        "exclusion_criteria": 40,
        "expertise_areas": 50,
        "primary_topics": 40,
        "audience_focus": 40,
        "aliases": 40,
        "typical_topics": 40,
        "email": 28,
        "team": 14,
        "industry": 22,
        "role_seniority": 16,
        "location_country": 16,
        "location_city": 20,
        "linkedin_url": 36,
        "github_url": 36,
        "website_url": 36,
        "homepage": 36,
        "slug": 22,
        "typical_month": 14,
        "is_active": 12,
        "pending_review": 16,
        "display_order": 16,
    }
    for idx, col in enumerate(spec.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col.name, 22)
