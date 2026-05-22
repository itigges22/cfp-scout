"""Workbook → typed rows (plan 31).

Validates each cell against its column spec, accumulating errors per row
without short-circuiting (the team wants to see EVERY problem in one pass,
not fix one at a time).

The returned ``ParsedWorkbook`` is consumed by diff.py to build the
inserts/updates/deletes list per sheet.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.services.workbook._cells import (
    CellError,
    parse_action,
    parse_bool,
    parse_date,
    parse_enum,
    parse_int,
    parse_list_text,
    parse_long_text,
    parse_text,
    parse_uuid,
)
from app.services.workbook._schema import SHEET_SPECS, ColumnSpec, SheetSpec


@dataclass(slots=True, frozen=True)
class SheetRowError:
    sheet: str
    row: int  # 1-based, matches openpyxl row numbers
    field: str
    value: Any
    message: str


@dataclass(slots=True)
class ParsedSheet:
    sheet: str
    rows: list[dict[str, Any]] = field(default_factory=list)
    errors: list[SheetRowError] = field(default_factory=list)


@dataclass(slots=True)
class ParsedWorkbook:
    sheets: dict[str, ParsedSheet] = field(default_factory=dict)
    unknown_sheets: list[str] = field(default_factory=list)
    file_errors: list[str] = field(default_factory=list)

    def total_errors(self) -> int:
        return sum(len(s.errors) for s in self.sheets.values()) + len(self.file_errors)


def parse_workbook(content: bytes) -> ParsedWorkbook:
    """Parse XLSX bytes → ParsedWorkbook.

    Never raises on parse failures — returns a ParsedWorkbook with the
    file-level error noted. Per-row / per-cell failures land in
    ``ParsedSheet.errors``.
    """
    result = ParsedWorkbook()
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        result.file_errors.append(f"could not open workbook: {exc}")
        return result

    visible_sheets = wb.sheetnames
    unknown = [
        s for s in visible_sheets if s not in {sp.name for sp in SHEET_SPECS} and s != "Reference"
    ]
    result.unknown_sheets = unknown

    for spec in SHEET_SPECS:
        if spec.name not in visible_sheets:
            # Missing-sheet is fine — caller treats it as "no rows for that entity".
            result.sheets[spec.name] = ParsedSheet(sheet=spec.name)
            continue
        ws = wb[spec.name]
        result.sheets[spec.name] = _parse_sheet(spec, ws)

    return result


def _parse_sheet(spec: SheetSpec, ws) -> ParsedSheet:
    ps = ParsedSheet(sheet=spec.name)

    rows_iter = ws.iter_rows(values_only=False)
    header_row = next(rows_iter, None)
    if header_row is None:
        return ps

    header_cells: list[Cell] = list(header_row)
    header_names = [str(c.value).strip() if c.value is not None else "" for c in header_cells]

    # Map column-spec name → its header column index (0-based). Missing
    # columns are tolerated; required ones produce an error per row when
    # absent.
    col_index: dict[str, int] = {}
    for idx, name in enumerate(header_names):
        if name in {c.name for c in spec.columns}:
            col_index[name] = idx

    for row_num, raw_row in enumerate(rows_iter, start=2):  # data starts at row 2
        # openpyxl yields a tuple of Cell objects in read_only mode.
        cells = list(raw_row)
        # Skip fully-empty rows (Google Sheets exports often have trailing blanks).
        if all(_cell_is_empty(c) for c in cells):
            continue

        row_data, row_errors = _parse_row(spec, cells, col_index, row_num)
        ps.rows.append(row_data)
        ps.errors.extend(row_errors)

    return ps


def _parse_row(
    spec: SheetSpec,
    cells: list[Cell],
    col_index: dict[str, int],
    row_num: int,
) -> tuple[dict[str, Any], list[SheetRowError]]:
    out: dict[str, Any] = {}
    errs: list[SheetRowError] = []

    # _action first — affects whether required-field validation runs.
    raw_action: Any = None
    if (idx := col_index.get("_action")) is not None and idx < len(cells):
        raw_action = cells[idx].value
    try:
        action = parse_action("_action", raw_action)
    except CellError as ce:
        errs.append(SheetRowError(spec.name, row_num, ce.field, ce.value, ce.message))
        action = "upsert"  # best-effort continue
    out["_action"] = action

    # _scout_id
    raw_id: Any = None
    if (idx := col_index.get("_scout_id")) is not None and idx < len(cells):
        raw_id = cells[idx].value
    try:
        out["_scout_id"] = parse_uuid("_scout_id", raw_id, required=False)
    except CellError as ce:
        errs.append(SheetRowError(spec.name, row_num, ce.field, ce.value, ce.message))
        out["_scout_id"] = None

    # Skip rows that are explicitly `skip`.
    if action == "skip":
        return out, errs

    # delete rows: skip required validation for non-id fields — we only need the id.
    skip_required = action == "delete"

    for col in spec.columns:
        if col.name in {"_action", "_scout_id"}:
            continue
        raw: Any = None
        if (idx := col_index.get(col.name)) is not None and idx < len(cells):
            raw = cells[idx].value
        required = col.required and not skip_required
        try:
            value = _parse_one(col, raw, required=required)
        except CellError as ce:
            errs.append(SheetRowError(spec.name, row_num, ce.field, ce.value, ce.message))
            value = None
        out[col.name] = value

    return out, errs


def _parse_one(col: ColumnSpec, raw: Any, *, required: bool) -> Any:
    if col.kind == "text":
        return parse_text(col.name, raw, required=required, max_len=col.max_len)
    if col.kind == "long_text":
        return parse_long_text(col.name, raw, required=required)
    if col.kind == "int":
        return parse_int(col.name, raw, required=required)
    if col.kind == "float":
        return parse_int(col.name, raw, required=required)  # treat float as int for our schema
    if col.kind == "bool":
        return parse_bool(col.name, raw, required=required)
    if col.kind == "date":
        return parse_date(col.name, raw, required=required)
    if col.kind == "uuid":
        return parse_uuid(col.name, raw, required=required)
    if col.kind == "list_text":
        return parse_list_text(col.name, raw, required=required)
    if col.kind == "list_uuid":
        # Shouldn't occur in pass-1 sheets — SMEs reference Topics/Audiences by NAME
        # and the apply layer resolves to UUIDs. Kept for future use.
        return parse_list_text(col.name, raw, required=required)
    if col.kind == "enum":
        return parse_enum(col.name, raw, required=required, allowed=col.enum_values)
    if col.kind == "action":
        return parse_action(col.name, raw)
    raise ValueError(f"unsupported column kind: {col.kind}")


def _cell_is_empty(cell: Cell) -> bool:
    return cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())
