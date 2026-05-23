"""Build the initial Scout config workbook from Red Hat AI source docs.

Output: ``rh-docs/scout-initial-config.xlsx`` (gitignored).

Run from the repo root::

  podman exec scout-api /app/.venv/bin/python /app/scripts/build_initial_workbook.py

The output file is then uploaded via Scout's UI:

  /settings → "Workbook import / export" → Upload & preview → Apply.

The script is a pure data-to-XLSX converter — no DB writes, no LLM calls.
All content tables live in this file so the user can edit them and re-run.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Tables — edit these to retune the workbook content.
# ---------------------------------------------------------------------------

PILLARS = [
    # Verbatim from the Red Hat AI Platform Customer Deck — these are the
    # four customer-facing pillars on the "Scaling AI across the hybrid
    # cloud" master slide. Cross-checked against the Q4 2026 messaging
    # guide where the same four show up as the "Why Red Hat AI" sections.
    {
        "name": "Flexible and efficient inferencing",
        "description": (
            "Fast, flexible, and cost-effective model deployments across a "
            "diverse footprint. vLLM for maximum throughput + minimum "
            "latency, llm-d for distributed inference, LLM compressor for "
            "reduced compute utilization, and the Red Hat AI repository on "
            "Hugging Face for pre-optimized models. Granite models, "
            "Apache-2.0-licensed and indemnified, are the canonical example."
        ),
        "display_order": 1,
    },
    {
        "name": "Connecting models to data",
        "description": (
            "Simplified, consistent experience to customize models with the "
            "organization's own data. Built-in data ingestion + "
            "pre-processing for unstructured sources (PDFs, docs), synthetic "
            "data generation when private data is thin, InstructLab for "
            "fine-tuning, RAG and RAFT patterns for grounded responses, and "
            "self-service IDEs (JupyterLab) for data scientists + AI "
            "engineers."
        ),
        "display_order": 2,
    },
    {
        "name": "Agentic AI innovation",
        "description": (
            "Agile, stable foundation to accelerate the development + "
            "deployment of AI agentic workflows. Built-in frameworks via "
            "ogx (previously Llama Stack), standardized communication "
            "protocols (MCP), the flexibility to integrate preferred tools "
            "(LangChain, Crew AI), agents as microservices, and the AI hub "
            "+ Gen AI studio dashboards for platform and AI engineers."
        ),
        "display_order": 3,
    },
    {
        "name": "Scaling AI across the hybrid cloud",
        "description": (
            "Enterprise-grade, flexible, and secure AI platform that builds, "
            "deploys, and manages AI models + agentic apps at scale across "
            "edge, private cloud, physical, virtual, and public cloud "
            "footprints. Private and sovereign AI practices, enhanced "
            "observability (platform metrics, zero-config GPU, AI "
            "performance metrics), on-prem + air-gapped deployment support."
        ),
        "display_order": 4,
    },
]

INDUSTRIES = [
    "Financial services",
    "Healthcare",
    "Government / Public sector",
    "Telecommunications",
    "Retail",
    "Manufacturing",
    "Energy",
    "Automotive",
    "Technology",
    "Education",
    "Media & entertainment",
]

AUDIENCES = [
    # The Brand and Audience insights Team's 11 canonical Red Hat personas,
    # transcribed from "Writing for Red Hat Notes". Goal / Content Needs /
    # Challenges / Solution-criteria all sourced verbatim and mapped into
    # Scout's audience_profiles shape.
    {
        "name": "C-Suite",
        "industry": "Technology",
        "role_seniority": "executive",
        "description": (
            "Helping their team and company grow. Buys based on business "
            "outcomes, ROI, and vendor stability — not features. Decides "
            "(or vetoes) at the strategic level; content has to lead with "
            "outcomes and proof of established reputation."
        ),
        "primary_pain_points": [
            "Financial concerns",
            "People: resources, alignment across groups, partners, competitors",
            "Replacing outdated or unsupported tech",
            "Reacting to changes from clients, partners, competitors",
        ],
        "key_messages": [
            "Cost / price efficiency",
            "Established vendor with strong reputation",
            "Proven ability to solve business issues",
            "Clear ROI or efficiency story",
        ],
    },
    {
        "name": "Line of Business (LOB)",
        "industry": "Technology",
        "role_seniority": "executive",
        "description": (
            "VP / Director of a business unit. Goal: satisfy customers. "
            "Buys when content informs strategic decisions, handles "
            "competitive threats, and meets financial objectives. "
            "Customer testimonials and reputation carry real weight."
        ),
        "primary_pain_points": [
            "Modernizing technology fast enough to stay ahead of competitors",
            "Legislation, especially around data security",
            "Changing market factors",
        ],
        "key_messages": [
            "Cost and ROI",
            "Customer testimonials",
            "Reputation",
            "Demonstrated impact on customer satisfaction",
        ],
    },
    {
        "name": "App Dev ITDM",
        "industry": "Technology",
        "role_seniority": "director",
        "description": (
            "Application Development IT Decision Maker — director or "
            "manager of a specialized dept. Buys when content addresses "
            "specific business needs, meets financial objectives, and "
            "meets client demand. Aligns the team's tooling with "
            "executive priorities."
        ),
        "primary_pain_points": [
            "Labor and skill shortages",
            "Alignment with executives",
            "Security",
            "Maintaining compliance with changing internal + external regulations",
        ],
        "key_messages": [
            "Cost or ROI",
            "Alignment with executives",
            "Security posture",
            "Stable, reliable solutions",
        ],
    },
    {
        "name": "I.T. Operations (Manager of IT)",
        "industry": "Technology",
        "role_seniority": "manager",
        "description": (
            "Manager of IT operations. Goal: having an impact on end "
            "users and the business. Buys when content addresses "
            "modernization, handles specific incidents (data breaches, "
            "competitive threats), and supports tech-sunsetting paths."
        ),
        "primary_pain_points": [
            "Security",
            "External influences",
            "Labor shortages",
            "Culture (resistance to change, alignment difficulty on initiatives)",
        ],
        "key_messages": [
            "Cost and overall value — support, services, functionality, training",
            "Business or industry knowledge",
            "Reputation",
            "Tested, proven solutions",
        ],
    },
    {
        "name": "Enterprise Architect",
        "industry": "Technology",
        "role_seniority": "director",
        "description": (
            "Goal: acquiring new skills and working with new tech. Buys "
            "when content helps them stay competitive and aware of new "
            "technologies, improves productivity, and provides a "
            "refresh / replace path for hardware or unsupported software."
        ),
        "primary_pain_points": [
            "Culture (change management, buy vs build)",
            "People (talent gaps, differing values)",
            "Impact on existing systems when implementing new technology",
        ],
        "key_messages": [
            "Proven knowledge of the industry or business",
            "Existing relationship with the company",
            "Established company with strong reputation",
            "Appealing support structure that meets business / team needs",
        ],
    },
    {
        "name": "Procurement",
        "industry": "Technology",
        "role_seniority": "manager",
        "description": (
            "Goal: meeting personal goals and tackling challenging "
            "tasks. Buys when content addresses external challenges "
            "(competitive threats, partner shifts, client expectations) "
            "and meets financial objectives."
        ),
        "primary_pain_points": [
            "External influences (buying cycle, profits, market changes, supply chain)",
            "Technology — automation, efficiency, supporting remote work",
        ],
        "key_messages": [
            "Cost and ROI",
            "Meets business needs",
            "Offers support for implementation and beyond",
            "Credible vendor with good reputation and expertise",
        ],
    },
    {
        "name": "Sys Admin",
        "industry": "Technology",
        "role_seniority": "ic",
        "description": (
            "Goal: overcoming challenges with new technologies and "
            "improving the lives of others. Buys when content meets "
            "specific business needs, modernization or innovation "
            "needs, and financial objectives."
        ),
        "primary_pain_points": [
            "Financial — staying under budget, cost reduction, budget limits",
            "Keeping up with technology changes and digital transformation",
            "Security (data, cloud, threat actors)",
        ],
        "key_messages": [
            "Cost-effectiveness",
            "Reliability under change",
            "Security-by-default posture",
            "Operator-friendly tooling",
        ],
    },
    {
        "name": "I.T. Security Practitioner",
        "industry": "Technology",
        "role_seniority": "director",
        "description": (
            "CISO or Head of Security. Goal: addressing new challenges "
            "with security tech, working with their team, making a "
            "difference at their company. Buys when content addresses "
            "audit gaps, replaces outdated tech, and meets specific "
            "client requirements."
        ),
        "primary_pain_points": [
            "Dynamic threat landscape",
            "Resource constraints",
            "Maintaining high security in a changing tech landscape",
            "Ensuring employees follow protocol",
        ],
        "key_messages": [
            "Pricing",
            "Appealing support structure",
            "Proven knowledge of industry or business",
            "Strong reputation; no recent breaches",
        ],
        "exclusion_criteria": [
            "Vendors with prior data breaches or major security incidents",
        ],
    },
    {
        "name": "Automation Architect",
        "industry": "Technology",
        "role_seniority": "ic",
        "description": (
            "Goal: using data analytics and figuring out creative "
            "automation solutions. Buys when content addresses urgent "
            "need (EOL, expired support, security crisis), demonstrates "
            "process improvement, supports strategic initiatives, and "
            "integrates easily with existing systems."
        ),
        "primary_pain_points": [
            "Staying up to date with automation + cloud trends",
            "Getting the right tools into employees' hands",
            "People (talent retention, hiring, resistance to remote, automation limits)",
        ],
        "key_messages": [
            "Vendor stability and reputation",
            "Understanding of business processes and needs",
            "Industry expertise",
            "Easy integration with existing systems",
        ],
    },
    {
        "name": "Data Scientist",
        "industry": "Technology",
        "role_seniority": "ic",
        "description": (
            "Goal: using data analytics to drive innovation and growth. "
            "Buys when content addresses strategic initiatives, "
            "overcomes obsolescence or operational inefficiencies, and "
            "addresses data-privacy / security concerns."
        ),
        "primary_pain_points": [
            "Staying current in a rapidly changing field",
            "Optimizing use of available data and best practices",
            "Financial constraints",
            "Getting data into a workable format (cleaning, combining sheets)",
        ],
        "key_messages": [
            "De-risks adopting new techniques",
            "Reduces data-prep burden",
            "Privacy + security guardrails built in",
            "Drives measurable business outcomes from data",
        ],
    },
    {
        "name": "Developer",
        "industry": "Technology",
        "role_seniority": "ic",
        "description": (
            "Senior / lead software engineer. Goal: developing "
            "solutions that are widely used. Buys when content "
            "addresses specific use cases, shows cost reduction, "
            "demonstrates obsolescence of current tooling, and proves "
            "the solution delivers."
        ),
        "primary_pain_points": [
            "Staying ahead amid a constantly changing tech landscape",
            "Communication (changes within ecosystem breaking products, fix prioritization)",
            "Resources (finding the right people with the right skills)",
            "Architectural challenges",
        ],
        "key_messages": [
            "Cost (discounts available)",
            "Reputation, credibility, and stability",
            "Solution meets specific business needs",
            "Strong developer experience and documentation",
        ],
    },
]

# Topic vocabulary — controlled. Aliases help extraction normalize variants.
TOPICS = [
    ("LLMs", ["large language models", "language models", "foundation models"]),
    ("RAG", ["retrieval augmented generation", "retrieval-augmented generation"]),
    ("Agents", ["AI agents", "agentic AI", "agentic workflows"]),
    ("Fine-tuning", ["LoRA", "PEFT", "instruction tuning", "InstructLab"]),
    ("MLOps", ["ML ops", "ML operations", "model ops"]),
    ("Inference", ["model inference", "model serving", "serving"]),
    ("vLLM", ["vllm"]),
    ("KServe", ["kserve", "kubeflow serving"]),
    ("GPU", ["GPUs", "accelerators", "NVIDIA", "AMD GPU"]),
    ("OpenShift", ["openshift", "OpenShift Container Platform", "OCP"]),
    ("Kubernetes", ["k8s", "k8"]),
    ("OpenShift AI", ["openshift ai", "RHOAI", "openshift-ai"]),
    ("RHEL AI", ["rhel ai", "Red Hat Enterprise Linux AI"]),
    ("AI Inference Server", ["ais", "red hat ai inference server"]),
    ("Red Hat AI Enterprise", ["RHAIE", "AI Enterprise"]),
    ("Granite", ["granite models", "ibm granite"]),
    ("Llama", ["meta llama", "llama 3", "llama 4"]),
    ("Mistral", ["mistral ai"]),
    ("Embeddings", ["text embeddings", "vector embeddings"]),
    ("Vector databases", ["vector db", "vector store", "pgvector", "milvus", "chroma"]),
    ("OpenAI API", ["openai compatible", "openai-compatible API"]),
    ("Prompt engineering", ["prompting", "system prompts"]),
    ("Evaluations", ["evals", "eval harness", "benchmarks"]),
    ("Guardrails", ["safety classifier", "content filtering", "AI safety"]),
    ("Observability", ["monitoring", "prometheus", "grafana", "OpenTelemetry"]),
    ("Hybrid cloud", ["multi-cloud", "edge"]),
    ("Edge AI", ["edge inference", "device AI"]),
    ("Confidential computing", ["confidential AI", "enclave"]),
    ("Model quantization", ["compression", "GPTQ", "AWQ", "INT8"]),
    ("PyTorch", ["pytorch", "torch"]),
    ("Hugging Face", ["huggingface", "hf"]),
    ("Distributed training", ["multi-GPU training", "FSDP", "DeepSpeed"]),
    ("Data pipelines", ["ETL", "feature engineering", "data ingestion"]),
    ("Cost optimization", ["TCO", "GPU utilization"]),
    ("Open source AI", ["open models", "open weights"]),
    ("InstructLab", ["lab tuning"]),
    ("Llama Stack", ["llama-stack"]),
    ("Docling", ["docling pdf"]),
]

# SMEs — first names only, from the metrics file. Bio is placeholder; user
# fills full name + email + bio + topic/audience assignments in the XLSX
# before uploading (or re-imports after editing in the SPA).
SMES = [
    {"full_name": "Cedric", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Sasa", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Addie", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Legare", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Sawyer", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Grace", "team": "DAAM", "expertise_areas": ["AI advocacy"]},
    {"full_name": "Taylor", "team": "TMM", "expertise_areas": ["Technical marketing"]},
]


# ---------------------------------------------------------------------------
# Workbook plumbing
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)


def _set_headers(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autosize(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        max_w = max(len(h), 12)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_w = max(max_w, min(len(str(v).split("\n")[0]), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_w + 2


def _semicolon_list(items: list[str]) -> str:
    return ";".join(items)


def write_pillars(ws) -> None:
    headers = ["_scout_id", "_action", "name", "description", "display_order"]
    _set_headers(ws, headers)
    for i, p in enumerate(PILLARS, start=2):
        ws.cell(row=i, column=1, value="")  # _scout_id blank = insert
        ws.cell(row=i, column=2, value="upsert")
        ws.cell(row=i, column=3, value=p["name"])
        ws.cell(row=i, column=4, value=p["description"])
        ws.cell(row=i, column=5, value=p["display_order"])
    _autosize(ws, headers)


def write_industries(ws) -> None:
    headers = ["_scout_id", "_action", "name"]
    _set_headers(ws, headers)
    for i, name in enumerate(INDUSTRIES, start=2):
        ws.cell(row=i, column=1, value="")
        ws.cell(row=i, column=2, value="upsert")
        ws.cell(row=i, column=3, value=name)
    _autosize(ws, headers)


def write_audiences(ws) -> None:
    headers = [
        "_scout_id",
        "_action",
        "name",
        "industry",
        "role_seniority",
        "description",
        "primary_pain_points",
        "key_messages",
        "exclusion_criteria",
        "is_active",
    ]
    _set_headers(ws, headers)
    for i, a in enumerate(AUDIENCES, start=2):
        ws.cell(row=i, column=1, value="")
        ws.cell(row=i, column=2, value="upsert")
        ws.cell(row=i, column=3, value=a["name"])
        ws.cell(row=i, column=4, value=a["industry"])
        ws.cell(row=i, column=5, value=a["role_seniority"])
        ws.cell(row=i, column=6, value=a["description"])
        ws.cell(row=i, column=7, value=_semicolon_list(a["primary_pain_points"]))
        ws.cell(row=i, column=8, value=_semicolon_list(a["key_messages"]))
        ws.cell(row=i, column=9, value="")
        ws.cell(row=i, column=10, value="TRUE")
    _autosize(ws, headers)


def write_smes(ws) -> None:
    headers = [
        "_scout_id",
        "_action",
        "full_name",
        "email",
        "team",
        "expertise_areas",
        "primary_topics",
        "audience_focus",
        "location_country",
        "location_city",
        "bio",
        "linkedin_url",
        "github_url",
        "website_url",
        "is_active",
    ]
    _set_headers(ws, headers)
    for i, s in enumerate(SMES, start=2):
        ws.cell(row=i, column=1, value="")
        ws.cell(row=i, column=2, value="upsert")
        ws.cell(row=i, column=3, value=s["full_name"])
        ws.cell(row=i, column=4, value="")  # email TBD
        ws.cell(row=i, column=5, value=s["team"])
        ws.cell(row=i, column=6, value=_semicolon_list(s["expertise_areas"]))
        ws.cell(row=i, column=7, value="")  # primary_topics TBD
        ws.cell(row=i, column=8, value="")  # audience_focus TBD
        ws.cell(row=i, column=9, value="US")  # placeholder country
        ws.cell(row=i, column=10, value="")  # city TBD
        ws.cell(
            row=i,
            column=11,
            value=(
                f"{s['full_name']} is a Red Hat {s['team']} subject-matter "
                f"expert in AI advocacy and developer engagement. Bio to be "
                f"completed by the team — minimum 200 characters required."
            ),
        )
        ws.cell(row=i, column=12, value="")
        ws.cell(row=i, column=13, value="")
        ws.cell(row=i, column=14, value="")
        ws.cell(row=i, column=15, value="TRUE")
    _autosize(ws, headers)


def write_topics(ws) -> None:
    headers = [
        "_scout_id",
        "_action",
        "name",
        "slug",
        "aliases",
        "is_active",
        "pending_review",
    ]
    _set_headers(ws, headers)
    for i, (name, aliases) in enumerate(TOPICS, start=2):
        ws.cell(row=i, column=1, value="")
        ws.cell(row=i, column=2, value="upsert")
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value="")  # let the apply layer derive slug
        ws.cell(row=i, column=5, value=_semicolon_list(aliases))
        ws.cell(row=i, column=6, value="TRUE")
        ws.cell(row=i, column=7, value="FALSE")
    _autosize(ws, headers)


def write_series(ws) -> None:
    # Empty series sheet — 35 are already seeded by the baseline migration.
    headers = [
        "_scout_id",
        "_action",
        "canonical_name",
        "aliases",
        "description",
        "typical_month",
        "typical_topics",
        "homepage",
        "is_active",
    ]
    _set_headers(ws, headers)
    _autosize(ws, headers)


def write_reference(ws) -> None:
    """Quick reference / instructions sheet."""
    ws["A1"] = "Scout initial-config workbook"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        "",
        "Generated by scripts/build_initial_workbook.py from rh-docs/.",
        "",
        "How to upload:",
        "  1. Open Scout → /settings → Workbook import / export.",
        "  2. Click 'Upload & preview…' and pick this file.",
        "  3. Review the per-sheet diff table.",
        "  4. Click 'Apply' to commit.",
        "",
        "Per-sheet notes:",
        "  • Pillars — 4 rows. Edit names + descriptions as needed.",
        "  • Industries — 11 rows. Standard set.",
        "  • Audiences — 8 personas with pain points + key messages.",
        "  • SMEs — 7 rows. First names only; complete bio + topics + audience_focus.",
        "  • Topics — ~37 controlled-vocab terms with aliases.",
        "  • Series — empty (35 already seeded by the baseline migration).",
        "",
        "Editing convention:",
        "  • _action=upsert (default) inserts new + updates existing by _scout_id.",
        "  • Leave _scout_id blank on new rows; the apply layer fills it.",
        "  • Semicolon-separated lists for arrays (e.g. 'a;b;c').",
        "  • Booleans: TRUE / FALSE (case-insensitive).",
    ]
    for i, line in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(out_path: Path) -> None:
    wb = Workbook()
    # Default sheet → rename to Reference
    ref = wb.active
    ref.title = "Reference"
    write_reference(ref)

    write_pillars(wb.create_sheet("Pillars"))
    write_industries(wb.create_sheet("Industries"))
    write_audiences(wb.create_sheet("Audiences"))
    write_smes(wb.create_sheet("SMEs"))
    write_topics(wb.create_sheet("Topics"))
    write_series(wb.create_sheet("Series"))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK → {out_path} ({out_path.stat().st_size:,} bytes)")
    print(
        "Counts:  "
        f"pillars={len(PILLARS)}  industries={len(INDUSTRIES)}  "
        f"audiences={len(AUDIENCES)}  smes={len(SMES)}  topics={len(TOPICS)}"
    )


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("rh-docs/scout-initial-config.xlsx")
    main(out)
